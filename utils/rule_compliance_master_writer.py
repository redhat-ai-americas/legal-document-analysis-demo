#!/usr/bin/env python3
"""
Rule Compliance Master Writer

Creates a master Excel workbook that aggregates rule compliance results across multiple documents.
Perfect for batch processing where you need to see compliance status for all documents at once.
"""

import os
import json
from datetime import datetime
from typing import Dict, Any, List

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    Workbook = None
    Worksheet = None


class RuleComplianceMasterWriter:
    """Manages a master Excel workbook for multiple document compliance results"""
    
    def __init__(self, master_path: str, rules_path: str):
        """
        Initialize the master writer
        
        Args:
            master_path: Path to master Excel file
            rules_path: Path to rules file for loading all rules
        """
        self.master_path = master_path
        self.rules_path = rules_path
        self.all_rules = self._load_all_rules()
        
    def _load_all_rules(self) -> List[Dict[str, Any]]:
        """Load all rules from file"""
        if not os.path.exists(self.rules_path):
            return []
        
        with open(self.rules_path, 'r') as f:
            if self.rules_path.endswith('.json'):
                data = json.load(f)
            elif self.rules_path.endswith(('.yaml', '.yml')):
                import yaml
                data = yaml.safe_load(f)
            else:
                return []
        
        return data.get('rules', [])
    
    def initialize_master_workbook(self) -> Workbook:
        """Create or load the master workbook with proper structure"""
        
        if os.path.exists(self.master_path):
            wb = openpyxl.load_workbook(self.master_path)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # Ensure required sheets exist
        if 'Compliance Matrix' not in wb.sheetnames:
            self._create_compliance_matrix_sheet(wb)
        
        if 'Document Summary' not in wb.sheetnames:
            self._create_document_summary_sheet(wb)
        
        if 'Critical Issues' not in wb.sheetnames:
            self._create_critical_issues_sheet(wb)
        
        if 'Rule Details' not in wb.sheetnames:
            self._create_rule_details_sheet(wb)
        
        return wb
    
    def _create_compliance_matrix_sheet(self, wb: Workbook) -> None:
        """Create the main compliance matrix showing all documents vs all rules"""
        
        ws = wb.create_sheet('Compliance Matrix', 0)
        
        # Title
        ws.merge_cells('A1:Z1')
        cell = ws['A1']
        cell.value = "Rule Compliance Matrix - All Documents"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
        
        # Headers will be set when first document is added
        # Row 2: empty (spacing)
        # Row 3: Column headers (Document, Overall Score, then each rule)
        # Row 4+: Document rows
        
    def _create_document_summary_sheet(self, wb: Workbook) -> None:
        """Create summary sheet showing overall metrics for each document"""
        
        ws = wb.create_sheet('Document Summary')
        
        # Title
        ws.merge_cells('A1:H1')
        cell = ws['A1']
        cell.value = "Document Compliance Summary"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = [
            "Document Name",
            "Analysis Date",
            "Overall Score",
            "Total Rules",
            "Compliant",
            "Non-Compliant",
            "Critical Issues",
            "Status"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.freeze_panes = 'A4'
    
    def _create_critical_issues_sheet(self, wb: Workbook) -> None:
        """Create sheet listing all critical and high-priority issues across documents"""
        
        ws = wb.create_sheet('Critical Issues')
        
        # Title
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = "Critical & High Priority Issues Across All Documents"
        cell.font = Font(bold=True, size=14, color='FF0000')
        cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = [
            "Document",
            "Rule Name",
            "Severity",
            "Status",
            "Issue Description",
            "Required Action"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.freeze_panes = 'A4'
    
    def _create_rule_details_sheet(self, wb: Workbook) -> None:
        """Create reference sheet with all rule definitions"""
        
        ws = wb.create_sheet('Rule Details')
        
        # Title
        ws.merge_cells('A1:E1')
        cell = ws['A1']
        cell.value = "Rule Definitions and Requirements"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = [
            "Rule ID",
            "Rule Name",
            "Severity",
            "Description",
            "Requirements"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add all rules
        current_row = 4
        for rule in self.all_rules:
            ws.cell(row=current_row, column=1, value=rule.get('id', ''))
            ws.cell(row=current_row, column=2, value=rule.get('name', ''))
            ws.cell(row=current_row, column=3, value=rule.get('severity', 'medium').upper())
            ws.cell(row=current_row, column=4, value=rule.get('description', ''))
            ws.cell(row=current_row, column=5, value=rule.get('rule_text', ''))
            
            # Color severity
            severity_cell = ws.cell(row=current_row, column=3)
            if rule.get('severity') == 'critical':
                severity_cell.font = Font(color='FF0000', bold=True)
            elif rule.get('severity') == 'high':
                severity_cell.font = Font(color='FF6600', bold=True)
            
            current_row += 1
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 60
        
        ws.freeze_panes = 'A4'
    
    def add_document_results(self, 
                            document_path: str,
                            analysis_data: Dict[str, Any]) -> None:
        """
        Add or update results for a document in the master workbook
        
        Args:
            document_path: Path to the analyzed document
            analysis_data: Complete analysis results including rule_compliance_results
        """
        
        if not EXCEL_AVAILABLE:
            print("❌ Excel update requires openpyxl library")
            return
        
        wb = self.initialize_master_workbook()
        doc_name = os.path.basename(document_path)
        
        # Update each sheet
        self._update_compliance_matrix(wb, doc_name, analysis_data)
        self._update_document_summary(wb, doc_name, analysis_data)
        self._update_critical_issues(wb, doc_name, analysis_data)
        
        # Save workbook
        os.makedirs(os.path.dirname(self.master_path), exist_ok=True)
        wb.save(self.master_path)
        
        print(f"✅ Updated master compliance workbook: {self.master_path}")
    
    def _update_compliance_matrix(self, wb: Workbook, doc_name: str, analysis_data: Dict[str, Any]) -> None:
        """Update the compliance matrix sheet"""
        
        ws = wb['Compliance Matrix']
        compliance_results = analysis_data.get('rule_compliance_results', [])
        summary = analysis_data.get('rule_compliance_summary', {})
        
        # Create headers if this is the first document
        if ws.cell(row=3, column=1).value != "Document":
            # Set up headers
            ws.cell(row=3, column=1, value="Document")
            ws.cell(row=3, column=2, value="Overall Score")
            
            # Add rule columns
            col_num = 3
            for rule in self.all_rules:
                cell = ws.cell(row=3, column=col_num)
                cell.value = rule.get('name', rule.get('id', 'Unknown'))
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, text_rotation=90)
                col_num += 1
            
            # Format header row
            for col in range(1, col_num):
                cell = ws.cell(row=3, column=col)
                if col <= 2:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Find or create row for this document
        doc_row = None
        for row_num in range(4, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == doc_name:
                doc_row = row_num
                break
        
        if doc_row is None:
            doc_row = ws.max_row + 1 if ws.max_row >= 4 else 4
        
        # Update document row
        ws.cell(row=doc_row, column=1, value=doc_name)
        
        # Overall score
        score = summary.get('compliance_score', 0) * 100
        score_cell = ws.cell(row=doc_row, column=2)
        score_cell.value = f"{score:.1f}%"
        
        # Color code the score
        if score < 60:
            score_cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            score_cell.font = Font(color='FF0000', bold=True)
        elif score < 80:
            score_cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            score_cell.font = Font(color='FF6600', bold=True)
        else:
            score_cell.fill = PatternFill(start_color='E6F3E6', end_color='E6F3E6', fill_type='solid')
            score_cell.font = Font(color='008000', bold=True)
        
        # Create results lookup
        results_by_id = {r['rule_id']: r for r in compliance_results}
        
        # Add compliance status for each rule
        col_num = 3
        for rule in self.all_rules:
            rule_id = rule.get('id')
            result = results_by_id.get(rule_id, {})
            
            if result:
                status = result.get('status', 'not_evaluated')
                # Use symbols for compact display
                status_symbol = {
                    'compliant': '✓',
                    'non_compliant': '✗',
                    'partially_compliant': '~',
                    'not_applicable': '-',
                    'requires_review': '?',
                    'not_evaluated': ''
                }.get(status, '')
                
                cell = ws.cell(row=doc_row, column=col_num)
                cell.value = status_symbol
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Color based on status
                if status == 'compliant':
                    cell.fill = PatternFill(start_color='E6F3E6', end_color='E6F3E6', fill_type='solid')
                    cell.font = Font(color='008000', bold=True)
                elif status == 'non_compliant':
                    cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                    cell.font = Font(color='FF0000', bold=True)
                elif status == 'partially_compliant':
                    cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                elif status == 'requires_review':
                    cell.fill = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')
            
            col_num += 1
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            if col_letter in ws.column_dimensions:
                ws.column_dimensions[col_letter].width = 5
        
        ws.freeze_panes = 'C4'
    
    def _update_document_summary(self, wb: Workbook, doc_name: str, analysis_data: Dict[str, Any]) -> None:
        """Update the document summary sheet"""
        
        ws = wb['Document Summary']
        summary = analysis_data.get('rule_compliance_summary', {})
        
        # Find or create row for this document
        doc_row = None
        for row_num in range(4, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == doc_name:
                doc_row = row_num
                break
        
        if doc_row is None:
            doc_row = ws.max_row + 1 if ws.max_row >= 4 else 4
        
        # Calculate overall status
        score = summary.get('compliance_score', 0) * 100
        if score < 60:
            status = "❌ High Risk"
        elif score < 80:
            status = "⚠️ Medium Risk"
        elif score < 95:
            status = "✓ Low Risk"
        else:
            status = "✅ Compliant"
        
        # Update row
        row_data = [
            doc_name,
            datetime.now().strftime('%Y-%m-%d %H:%M'),
            f"{score:.1f}%",
            summary.get('total_rules', 0),
            summary.get('compliant', 0),
            summary.get('non_compliant', 0),
            summary.get('critical_issues', 0),
            status
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=doc_row, column=col_num)
            cell.value = value
            cell.alignment = Alignment(vertical='center')
            
            # Highlight critical issues
            if col_num == 7 and value > 0:  # Critical issues column
                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                cell.font = Font(color='FF0000', bold=True)
            elif col_num == 8:  # Status column
                if "High Risk" in str(value):
                    cell.font = Font(color='FF0000', bold=True)
                elif "Medium Risk" in str(value):
                    cell.font = Font(color='FF6600', bold=True)
                else:
                    cell.font = Font(color='008000', bold=True)
        
        # Auto-adjust columns
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col_letter].width = 18
        ws.column_dimensions['A'].width = 30  # Document name
    
    def _update_critical_issues(self, wb: Workbook, doc_name: str, analysis_data: Dict[str, Any]) -> None:
        """Update the critical issues sheet"""
        
        ws = wb['Critical Issues']
        compliance_results = analysis_data.get('rule_compliance_results', [])
        
        # Filter for critical and high-priority non-compliant issues
        critical_issues = [
            r for r in compliance_results
            if r.get('status') in ['non_compliant', 'partially_compliant']
            and r.get('severity') in ['critical', 'high']
        ]
        
        # Remove existing entries for this document
        rows_to_delete = []
        for row_num in range(4, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == doc_name:
                rows_to_delete.append(row_num)
        
        # Delete in reverse order to maintain row numbers
        for row_num in reversed(rows_to_delete):
            ws.delete_rows(row_num, 1)
        
        # Add new issues
        if critical_issues:
            current_row = ws.max_row + 1 if ws.max_row >= 4 else 4
            
            for issue in critical_issues:
                # Determine action
                if issue.get('severity') == 'critical':
                    action = "Immediate legal review required"
                else:
                    action = "Review and negotiate amendments"
                
                row_data = [
                    doc_name,
                    issue.get('rule_name', issue.get('rule_id', '')),
                    issue.get('severity', '').upper(),
                    issue.get('status', '').replace('_', ' ').title(),
                    issue.get('rationale', '')[:200] + '...' if len(issue.get('rationale', '')) > 200 else issue.get('rationale', ''),
                    action
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = value
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    # Highlight critical severity
                    if col_num == 3 and value == 'CRITICAL':
                        cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                        cell.font = Font(color='FF0000', bold=True)
                
                current_row += 1
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 35


def update_master_compliance_excel(analysis_data: Dict[str, Any],
                                  target_document_path: str,
                                  rules_path: str,
                                  master_excel_path: str) -> str:
    """
    Update the master compliance Excel with results from a document
    
    Args:
        analysis_data: Complete analysis results
        target_document_path: Path to analyzed document
        rules_path: Path to rules file
        master_excel_path: Path to master Excel file
        
    Returns:
        Path to updated master Excel file
    """
    
    writer = RuleComplianceMasterWriter(master_excel_path, rules_path)
    writer.add_document_results(target_document_path, analysis_data)
    return master_excel_path