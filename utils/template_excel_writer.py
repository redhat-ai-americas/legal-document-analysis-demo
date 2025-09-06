#!/usr/bin/env python3
"""
Template Excel Writer

Creates Excel output that matches the Due Diligence Review Template format exactly.
This replaces the generic spreadsheet output with the specific template structure.
"""

import os
from datetime import datetime
from typing import Dict, Any, List, Tuple, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    # Define dummy types for when openpyxl is not available
    Workbook = None
    Worksheet = None

def extract_y_n_and_details(analysis_answer: str) -> Tuple[str, str]:
    """
    Extract Y/N determination and details from analysis answer.
    
    Args:
        analysis_answer: The LLM response text
        
    Returns:
        Tuple of (Y/N, details)
    """
    if not analysis_answer or analysis_answer.strip() in ['', 'MANUAL_FIELD', 'Not specified']:
        return 'N/A', 'Not analyzed'
    
    answer_lower = analysis_answer.lower()
    
    # Look for explicit Y/N indicators
    if any(indicator in answer_lower for indicator in ['no explicit mention', 'not specified', 'does not', 'no provisions']):
        return 'N', analysis_answer[:500] + '...' if len(analysis_answer) > 500 else analysis_answer
    elif any(indicator in answer_lower for indicator in ['yes', 'does have', 'includes', 'contains provisions']):
        return 'Y', analysis_answer[:500] + '...' if len(analysis_answer) > 500 else analysis_answer
    else:
        # Default to requiring review if unclear
        return 'Review', analysis_answer[:500] + '...' if len(analysis_answer) > 500 else analysis_answer

def _coerce_state_like(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Accept either the compact analysis_data dict or a full workflow state and
    coerce to an analysis-like dict shape for downstream mapping.
    """
    # If it looks like a full state with nested analysis
    if 'analysis_data' in data and isinstance(data['analysis_data'], dict):
        merged = {**data['analysis_data']}
        # Surface top-level rule results/entities if present on state
        if 'rule_compliance_results' in data and 'rule_compliance_results' not in merged:
            merged['rule_compliance_results'] = data.get('rule_compliance_results')
        if 'extracted_entities' in data and 'extracted_entities' not in merged:
            merged['extracted_entities'] = data.get('extracted_entities')
        return merged
    return data


def _shorten(text: str, max_len: int = 500) -> str:
    if not text:
        return ''
    return (text[: max_len] + '...') if len(text) > max_len else text


def _extract_company_names(extracted_entities: Optional[Dict[str, Any]]) -> Tuple[str, str]:
    company1 = 'Not found'
    company2 = 'Not found'
    if not extracted_entities:
        return company1, company2
    party1 = extracted_entities.get('party1')
    party2 = extracted_entities.get('party2')
    if isinstance(party1, dict) and party1.get('name'):
        company1 = party1.get('name')
    elif isinstance(extracted_entities.get('companies'), list) and extracted_entities['companies']:
        company1 = str(extracted_entities['companies'][0])
    if isinstance(party2, dict) and party2.get('name'):
        company2 = party2.get('name')
    elif isinstance(extracted_entities.get('companies'), list) and len(extracted_entities['companies']) > 1:
        company2 = str(extracted_entities['companies'][1])
    return company1, company2


def _friendly_rule_name(rule_entry: Dict[str, Any]) -> str:
    raw = str(rule_entry.get('rule_id') or rule_entry.get('name') or 'Rule').strip()
    name = raw.replace('_', ' ').strip().title()
    # Lightweight normalizations for common legal terms
    name = name.replace('Ip', 'IP').replace('Mfn', 'MFN').replace('Mfc', 'MFC').replace('Coc', 'CoC')
    name = name.replace('Gdpr', 'GDPR')
    return name


def _rules_to_columns(data: Dict[str, Any]) -> Tuple[Dict[str, str], List[str]]:
    """
    Build per-rule columns from rule_compliance_results.
    Returns (row_updates, rule_headers)
    """
    rule_results = data.get('rule_compliance_results') or []
    row_updates: Dict[str, str] = {}
    headers: List[str] = []
    seen_headers: set = set()
    for r in rule_results:
        name_str = _friendly_rule_name(r)
        status = str(r.get('status', 'unknown'))
        rationale = _shorten(str(r.get('rationale', '')), 1000)
        base_status = f"Rule: {name_str} (Status)"
        base_rationale = f"Rule: {name_str} (Rationale)"
        status_key = base_status
        rationale_key = base_rationale
        # Ensure uniqueness if duplicate friendly names
        suffix = 2
        while status_key in seen_headers or rationale_key in seen_headers:
            status_key = f"{base_status} #{suffix}"
            rationale_key = f"{base_rationale} #{suffix}"
            suffix += 1
        seen_headers.add(status_key)
        seen_headers.add(rationale_key)
        row_updates[status_key] = status
        row_updates[rationale_key] = rationale
        headers.extend([status_key, rationale_key])
    return row_updates, headers


def map_questionnaire_to_template_row(analysis_data: Dict[str, Any], 
                                    target_document_path: str, 
                                    reference_document_path: str) -> Dict[str, str]:
    """
    Map our questionnaire analysis data to the template row format.
    
    Args:
        analysis_data: Complete analysis results
        target_document_path: Path to analyzed document
        reference_document_path: Path to reference document
        
    Returns:
        Dictionary with template column values
    """
    data = _coerce_state_like(analysis_data)

    # Extract data sections
    doc_info = data.get('document_information', {})
    contract_summary = data.get('contract_summary', {})
    gaps = data.get('gaps_and_comments', {})
    exec_summary = data.get('executive_summary', {})
    
    # Helper function to safely get answer from question data
    def get_answer(section_data: Dict, question_id: str) -> str:
        question_data = section_data.get(question_id, {})
        if isinstance(question_data, dict):
            return question_data.get('answer', 'Not specified')
        return str(question_data) if question_data else 'Not specified'
    
    # Basic information columns
    row_data = {
        'Reviewer Name': get_answer(doc_info, 'reviewer_name'),
        'Target Name': get_answer(doc_info, 'target_company_name'),
        'Datasite Location': get_answer(doc_info, 'datasite_location'),
        'Document Name': get_answer(doc_info, 'document_name'),
        'Knowledge Gaps': get_answer(gaps, 'knowledge_gaps'),
        # Normalize contract start date to ISO if available from extracted_entities
        'Contract Start Date': (
            (data.get('extracted_entities', {}) or {}).get('contract_start_date', {}).get('standardized_date')
            if isinstance(data.get('extracted_entities'), dict) and data.get('extracted_entities', {}).get('contract_start_date')
            else get_answer(doc_info, 'contract_start_date').split('(')[0].strip()
        ),
        'Type of Agreement': get_answer(doc_info, 'agreement_type'),
    }

    # Add document-level metadata from extracted entities
    extracted_entities = data.get('extracted_entities', {})
    company1, company2 = _extract_company_names(extracted_entities)
    row_data['Company 1'] = company1
    row_data['Company 2'] = company2
    
    # Extract Y/N + Details from rule compliance results
    rule_results = data.get('rule_compliance_results', [])
    
    # Helper to find rule result by topic
    def get_rule_result(topic_keywords: List[str]) -> Tuple[str, str]:
        for rule in rule_results:
            rule_id = str(rule.get('rule_id', '') or rule.get('name', '')).lower()
            if any(kw in rule_id for kw in topic_keywords):
                status = rule.get('status', 'unknown')
                rationale = rule.get('rationale', '')
                # Map status to Y/N
                if status == 'non_compliant':
                    return 'Y', rationale  # Y = has issues
                elif status == 'compliant':
                    return 'N', 'No issues found'
                elif status == 'not_applicable':
                    return 'N/A', 'Not applicable'
                else:
                    return 'Review', rationale
        return 'N/A', 'Not evaluated by rules'
    
    # Source Code
    source_y_n, source_details = get_rule_result(['source_code', 'source code', 'escrow'])
    row_data['Source Code Y/N'] = source_y_n
    row_data['Source Code Details'] = source_details
    
    # Exclusivity/Non-Competes
    excl_y_n, excl_details = get_rule_result(['exclusiv', 'non_compet', 'non-compet'])
    row_data['Exclusivity Y/N'] = excl_y_n
    row_data['Exclusivity Details'] = excl_details
    
    # Forced Pricing Adjustment
    pricing_y_n, pricing_details = get_rule_result(['pricing', 'mfn', 'mfc', 'most_favored', 'price'])
    row_data['Pricing Y/N'] = pricing_y_n
    row_data['Pricing Details'] = pricing_details
    
    # IP Issues
    ip_y_n, ip_details = get_rule_result(['ip', 'intellectual', 'property', 'ownership'])
    row_data['IP Y/N'] = ip_y_n
    row_data['IP Details'] = ip_details
    
    # Limitation of Liability
    liability_y_n, liability_details = get_rule_result(['liability', 'limitation', 'damages', 'cap'])
    row_data['Liability Y/N'] = liability_y_n
    row_data['Liability Details'] = liability_details
    
    # Assignment/Change of Control
    assignment_y_n, assignment_details = get_rule_result(['assignment', 'assign', 'coc', 'change_of_control', 'change of control'])
    row_data['Assignment Y/N'] = assignment_y_n
    row_data['Assignment Details'] = assignment_details
    
    # Secondary Use Rights (not in our questionnaire, but template expects it)
    row_data['Secondary Use Rights'] = 'Not analyzed in current questionnaire'
    
    # Other material terms from rules
    indemnity_y_n, indemnity_details = get_rule_result(['indemnif', 'indemnity', 'indemnities'])
    other_terms = []
    if indemnity_y_n == 'Y' and indemnity_details:
        other_terms.append(f"Indemnity: {indemnity_details[:100]}...")
    
    # Add risk flags from analysis
    risk_assessment = data.get('risk_assessment', {})
    red_flags = risk_assessment.get('red_flag_summary', {}).get('total_red_flags', 0)
    if red_flags > 0:
        other_terms.append(f"{red_flags} risk flag(s) detected")
    
    row_data['Other Terms'] = '; '.join(other_terms) if other_terms else 'None identified'
    
    # Comments (internal comments + quality info)
    comments = []
    internal_comments = get_answer(gaps, 'internal_comments')
    if internal_comments and internal_comments != 'MANUAL_FIELD':
        comments.append(internal_comments)
    
    # Add quality information
    quality_metrics = data.get('quality_metrics', {})
    quality_grade = quality_metrics.get('quality_grade', 'Unknown')
    confidence = quality_metrics.get('overall_quality_score', 0)
    comments.append(f"Analysis Quality: {quality_grade} ({confidence:.0f}%)")
    
    # Add recommendation
    recommendation = exec_summary.get('overall_recommendation', '')
    if recommendation:
        comments.append(f"Recommendation: {recommendation}")
    
    row_data['Comments'] = '; '.join(comments)
    
    # Add per-rule columns
    rule_cols, _ = _rules_to_columns(data)
    row_data.update(rule_cols)

    return row_data

def create_template_excel(analysis_data: Dict[str, Any], 
                         output_path: str,
                         target_document_path: str, 
                         reference_document_path: str) -> str:
    """
    Create an Excel file matching the Due Diligence Review Template format.
    
    Args:
        analysis_data: Complete analysis results
        output_path: Path where Excel file should be saved
        target_document_path: Path to analyzed document
        reference_document_path: Path to reference document
        
    Returns:
        Path to created Excel file
    """
    if not EXCEL_AVAILABLE or openpyxl is None:
        print("❌ Excel creation requires openpyxl library")
        return ""
    
    try:
        # Generate Excel path
        excel_path = output_path.replace('.yaml', '_template.xlsx').replace('.yml', '_template.xlsx')
        
        # Map our data to template row format
        row_data = map_questionnaire_to_template_row(
            analysis_data, target_document_path, reference_document_path
        )
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Contract Analysis"
            
            # Define the base column structure (original) + metadata
            base_headers = [
                "Reviewer Name",
                "Target Name", 
                "Datasite Location",
                "Document Name",
                "Knowledge Gaps",
                "Contract Start Date",
                "Type of Agreement",
                "Company 1",
                "Company 2",
                "Source Code Y/N",
                "Source Code Details", 
                "Exclusivity Y/N",
                "Exclusivity Details",
                "Pricing Y/N", 
                "Pricing Details",
                "IP Y/N",
                "IP Details",
                "Liability Y/N", 
                "Liability Details",
                "Assignment Y/N",
                "Assignment Details",
                "Secondary Use Rights",
                "Other Terms",
                "Comments"
            ]

            # Add dynamic rule headers
            rule_headers = [h for h in row_data.keys() if h.startswith("Rule:")]
            headers = base_headers + rule_headers
            
            # Create header row with formatting
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                if cell is not None:
                    cell.value = header
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'), 
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Add data row
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num)
                if cell is not None:
                    # Map header to our row_data keys
                    value = row_data.get(header, 'Not specified')
                    # Add symbol for LLM-answerable key clause fields
                    if header in [
                        "Source Code Details", "Exclusivity Details", "Pricing Details",
                        "IP Details", "Liability Details", "Assignment Details"
                    ] and isinstance(value, str) and value not in ("Not specified", "MANUAL_FIELD"):
                        value = f"★ {value}"
                    cell.value = value
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                first_cell = column[0]
                if first_cell and first_cell.column is not None:
                    column_letter = get_column_letter(first_cell.column)
                    for cell in column:
                        try:
                            if cell.value is not None and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (ValueError, TypeError, AttributeError):
                            pass
                    adjusted_width = min(max_length + 2, 80)  # Wider cap to reduce truncation in view
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Add conditional formatting for Y/N columns
            # Adjusted indices due to added Company columns: Y/N columns now shifted by +2
            y_n_columns = [10, 12, 14, 16, 18, 20]  # Source, Exclusivity, Pricing, IP, Liability, Assignment
            for col_num in y_n_columns:
                cell = ws.cell(row=2, column=col_num)
                if cell is not None and cell.value:
                    if cell.value == 'Y':
                        cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')  # Light red
                    elif cell.value == 'N':
                        cell.fill = PatternFill(start_color='E6F3E6', end_color='E6F3E6', fill_type='solid')  # Light green
                    elif cell.value == 'Review':
                        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light yellow
        
        # Save workbook
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        wb.save(excel_path)
        
        print(f"✅ Template Excel created: {excel_path}")
        return excel_path
        
    except Exception as e:
        print(f"❌ Error creating template Excel: {e}")
        return ""

def update_master_template_excel(analysis_data: Dict[str, Any],
                               run_id: str,
                               target_document_path: str, 
                               reference_document_path: str,
                               master_excel_path: str) -> str:
    """
    Update or create a master Excel file in template format with multiple document rows.
    
    Args:
        analysis_data: Complete analysis results
        run_id: Run identifier
        target_document_path: Path to analyzed document 
        reference_document_path: Path to reference document
        master_excel_path: Path to master Excel file
        
    Returns:
        Path to updated master Excel file
    """
    if not EXCEL_AVAILABLE or openpyxl is None:
        print("❌ Excel update requires openpyxl library")
        return ""
    
    try:
        # Map our data to template row format
        row_data = map_questionnaire_to_template_row(
            analysis_data, target_document_path, reference_document_path
        )
        
        # Add run tracking
        row_data['Run ID'] = run_id
        row_data['Analysis Date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Define base headers (including tracking columns)
        base_headers = [
            "Run ID", "Analysis Date", "Reviewer Name", "Target Name", "Datasite Location",
            "Document Name", "Knowledge Gaps", "Contract Start Date", "Type of Agreement",
            "Company 1", "Company 2",
            "Source Code Y/N", "Source Code Details", "Exclusivity Y/N", "Exclusivity Details",
            "Pricing Y/N", "Pricing Details", "IP Y/N", "IP Details",
            "Liability Y/N", "Liability Details", "Assignment Y/N", "Assignment Details", 
            "Secondary Use Rights", "Other Terms", "Comments"
        ]

        # Rule headers for this run
        current_rule_headers = [h for h in row_data.keys() if h.startswith("Rule:")]
        
        # Check if master file exists
        if os.path.exists(master_excel_path):
            # Load existing workbook
            wb = openpyxl.load_workbook(master_excel_path)
            ws = wb.active
            
            if ws is not None:
                # Find document identifier to check for updates
                doc_name = os.path.basename(target_document_path)
                
                # Look for existing row with same document
                existing_row = None
                for row_num in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=6)  # Document Name column
                    if cell is not None and cell.value == doc_name:
                        existing_row = row_num
                        break
                
                if existing_row:
                    # Update existing row
                    target_row = existing_row
                    print(f"📊 Updating existing row {target_row} for document: {doc_name}")
                else:
                    # Add new row
                    target_row = ws.max_row + 1
                    print(f"📊 Adding new row {target_row} for document: {doc_name}")

                # Union headers: keep existing order, append any missing from base and rules
                existing_headers = [c.value for c in ws[1] if c.value]
                desired_headers = existing_headers[:]
                for h in base_headers:
                    if h not in desired_headers:
                        desired_headers.append(h)
                for h in current_rule_headers:
                    if h not in desired_headers:
                        desired_headers.append(h)

                # If headers changed, rewrite header row formatting for new columns
                if len(desired_headers) != len(existing_headers):
                    for idx, header in enumerate(desired_headers, 1):
                        cell = ws.cell(row=1, column=idx)
                        cell.value = header
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                print("❌ Error: Could not access worksheet")
                return ""
        else:
            # Create new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is not None:
                ws.title = "Contract Analysis Comparison"
                
                # Create headers (base only initially; rule headers added implicitly via row write below)
                desired_headers = base_headers + current_rule_headers
                for col_num, header in enumerate(desired_headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    if cell is not None:
                        cell.value = header
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                target_row = 2
                print(f"📊 Created new master Excel with document: {os.path.basename(target_document_path)}")
            else:
                print("❌ Error: Could not create worksheet")
                return ""
        
        # Determine final header set (row 1 values)
        headers = [c.value for c in ws[1] if c.value]

        # Ensure all rule headers present for this row; if any missing, append columns now
        for h in current_rule_headers:
            if h not in headers:
                headers.append(h)
                col_idx = len(headers)
                c = ws.cell(row=1, column=col_idx)
                c.value = h
                c.font = Font(bold=True, color='FFFFFF')
                c.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write data to target row
        if ws is not None:
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=target_row, column=col_num)
                if cell is not None:
                    cell.value = row_data.get(header, 'Not specified')
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Apply conditional formatting to Y/N columns in new/updated row
            # Adjusted for Company columns inserted after Type of Agreement
            # Base headers count up to Type of Agreement includes Company 1/2
            y_n_columns = [12, 14, 16, 18, 20, 22]
            for col_num in y_n_columns:
                cell = ws.cell(row=target_row, column=col_num)
                if cell is not None and cell.value:
                    if cell.value == 'Y':
                        cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                    elif cell.value == 'N':
                        cell.fill = PatternFill(start_color='E6F3E6', end_color='E6F3E6', fill_type='solid')
                    elif cell.value == 'Review':
                        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                first_cell = column[0]
                if first_cell and first_cell.column is not None:
                    column_letter = get_column_letter(first_cell.column)
                    for cell in column:
                        try:
                            if cell.value is not None and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (ValueError, TypeError, AttributeError):
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
        
        # Update or create LLM QA sheet for auditability
        try:
            qa_ws = wb["LLM QA"] if "LLM QA" in wb.sheetnames else wb.create_sheet("LLM QA")
            if qa_ws.max_row == 1 and (qa_ws.cell(row=1, column=1).value is None):
                # Write header
                qa_headers = [
                    "Run ID", "Document Name", "Section", "Field_ID", "Question",
                    "Answer_Source", "Confidence", "Answer"
                ]
                for col_num, h in enumerate(qa_headers, 1):
                    qa_ws.cell(row=1, column=col_num, value=h)
            # Append rows from analysis_data
            doc_name = os.path.basename(target_document_path)
            def emit_rows(section_key: str, pretty_section: str):
                section = analysis_data.get(section_key, {}) or {}
                for field_id, field in section.items():
                    if not isinstance(field, dict):
                        continue
                    question = field.get('question', field_id)
                    answer = field.get('answer', '')
                    confidence = field.get('confidence', None)
                    # Infer answer source
                    if section_key == 'document_information':
                        src = 'Deterministic'
                    elif section_key == 'gaps_and_comments':
                        src = 'Manual' if str(answer) == 'MANUAL_FIELD' else 'LLM'
                    else:
                        src = 'LLM'
                    qa_ws.append([
                        run_id, doc_name, pretty_section, field_id,
                        question, src, confidence, str(answer)[:2000]
                    ])
            emit_rows('document_information', 'Document Information')
            emit_rows('contract_summary', 'Contract Summary')
            emit_rows('gaps_and_comments', 'Gaps and Comments')
        except Exception:
            # Non-fatal
            pass

        # Save workbook
        os.makedirs(os.path.dirname(master_excel_path), exist_ok=True)
        wb.save(master_excel_path)
        
        print(f"📊 Master template Excel updated: {master_excel_path}")
        return master_excel_path
        
    except Exception as e:
        print(f"❌ Error updating master template Excel: {e}")
        return "" 

def _update_due_diligence_sheet(
    wb: "Workbook",
    analysis_data: Dict[str, Any],
    run_id: str,
    target_document_path: str
) -> None:
    """
    Create or update a 'Due Diligence' sheet with a two-row header:
    Row 1: Topic (e.g., Source Code)
    Row 2: Field (Any Issues? Y or N) and (Describe deviation)

    Populates values using rules compliance results and LLM answers.
    """
    try:
        ws_name = "Due Diligence"
        ws = wb[ws_name] if ws_name in wb.sheetnames else wb.create_sheet(ws_name)

        # Build two-row headers once
        if ws.max_row == 1 and (ws.cell(row=1, column=1).value is None):
            # Meta columns (single-row merged design simplified to two rows)
            meta_cols = [
                ("Meta", "Reviewer Name"),
                ("Meta", "Target Name"),
                ("Meta", "Customer name"),
                ("Meta", "Datasite Location"),
                ("Meta", "Document Name"),
                ("Meta", "Knowledge Gaps"),
                ("Meta", "Contract Start Date"),
                ("Meta", "Type of Agreement"),
            ]
            topics = [
                "Source Code",
                "Exclusivity/Non-Competes",
                "Forced Pricing Adjustment",
                "IP Issues",
                "Limitation of Liability",
                "Consent to Assign/CoC",
                "Indemnity",
                "Other",
                "Comments"
            ]
            # Compose header rows
            top_row: List[str] = []
            sub_row: List[str] = []
            for grp, field in meta_cols:
                top_row.append(grp)
                sub_row.append(field)
            for topic in topics:
                top_row.extend([topic, topic])
                sub_row.extend(["Any Issues? Y or N", "Describe deviation"])
            # Write headers
            for col, val in enumerate(top_row, 1):
                c = ws.cell(row=1, column=col, value=val)
                c.font = Font(bold=True, color='FFFFFF')
                c.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for col, val in enumerate(sub_row, 1):
                c = ws.cell(row=2, column=col, value=val)
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.freeze_panes = 'A3'

        # Build row values
        doc_name = os.path.basename(target_document_path)
        doc_info = analysis_data.get('document_information', {})
        gaps = analysis_data.get('gaps_and_comments', {})
        extracted_entities = analysis_data.get('extracted_entities', {})

        def gi(section: Dict[str, Any], fid: str, default: str = "") -> str:
            f = section.get(fid, {})
            return f.get('answer', default) if isinstance(f, dict) else default

        reviewer = gi(doc_info, 'reviewer_name', 'Automated Analysis')
        target_name = gi(doc_info, 'target_company_name', '')
        customer_name = gi(doc_info, 'counterparty_name', '')
        datasite = gi(doc_info, 'datasite_location', '')
        knowledge_gaps = gi(gaps, 'knowledge_gaps', '')
        start_date = (
            (extracted_entities.get('contract_start_date', {}) or {}).get('standardized_date')
            if isinstance(extracted_entities, dict) and extracted_entities.get('contract_start_date')
            else gi(doc_info, 'contract_start_date', '')
        )
        agreement_type = gi(doc_info, 'agreement_type', '')

        meta_values = [
            reviewer, target_name, customer_name, datasite, doc_name,
            knowledge_gaps, start_date, agreement_type
        ]

        # Map topics to questionnaire IDs for fallback descriptions
        topic_to_qid = {
            "Source Code": 'source_code_access',
            "Exclusivity/Non-Competes": 'exclusivity_non_competes',
            "Forced Pricing Adjustment": 'forced_pricing_adjustments',
            "IP Issues": 'ip_rights',
            "Limitation of Liability": 'limitation_of_liability',
            "Consent to Assign/CoC": 'assignment_coc',
            "Indemnity": 'indemnity',
            "Other": 'internal_comments',
            "Comments": 'internal_comments'
        }
        contract_summary = analysis_data.get('contract_summary', {}) or {}
        # Rules compliance summary by topic name
        rule_results = analysis_data.get('rule_compliance_results') or []
        def topic_summary(topic: str) -> (str, str):
            # Any issues: 'Y' if any non_compliant for matching topic name; else 'N' if any others; else ''
            relevant = [r for r in rule_results if (str(r.get('severity') or '') or True) and str(r.get('rule_id') or r.get('name') or '').lower().startswith(topic.split()[0].lower())]
            any_non = any((r.get('status') == 'non_compliant') for r in relevant)
            any_any = bool(relevant)
            issues = 'Y' if any_non else ('N' if any_any else '')
            # Describe deviation: join rationales for non_compliant, else LLM answer
            if any_non:
                rats = [str(r.get('rationale', '')) for r in relevant if r.get('status') == 'non_compliant']
                desc = ' '.join(r for r in rats if r)[:1000]
                if desc:
                    desc = f"★ {desc}"
            else:
                qid = topic_to_qid.get(topic)
                item = contract_summary.get(qid, {}) if isinstance(contract_summary, dict) else {}
                ans = item.get('answer', '') if isinstance(item, dict) else ''
                conf = item.get('confidence', None)
                if ans:
                    desc = f"★ {ans}"
                    if isinstance(conf, (int, float)):
                        desc = f"{desc} (conf {conf:.0f})"
                else:
                    desc = ''
            return issues, desc

        topics = [
            "Source Code",
            "Exclusivity/Non-Competes",
            "Forced Pricing Adjustment",
            "IP Issues",
            "Limitation of Liability",
            "Consent to Assign/CoC",
            "Indemnity",
            "Other",
            "Comments"
        ]
        topic_values: List[str] = []
        for t in topics:
            y_n, deviation = topic_summary(t)
            topic_values.extend([y_n, deviation])

        row_values = meta_values + topic_values

        # Find existing row by Document Name (5th column)
        target_row = None
        for r in range(3, ws.max_row + 1):
            if ws.cell(row=r, column=5).value == doc_name:
                target_row = r
                break
        if target_row is None:
            target_row = ws.max_row + 1
        for i, v in enumerate(row_values, 1):
            ws.cell(row=target_row, column=i, value=v)

        # Auto-width
        for col in ws.columns:
            try:
                max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
            except Exception:
                pass
    except Exception:
        pass

def reset_master_template_excel(master_excel_path: str, preserve_headers: bool = True) -> str:
    """
    Reset the master template Excel by removing all data rows while preserving headers,
    or recreating a new workbook with base headers if none exist.

    Args:
        master_excel_path: Path to the master Excel file to reset
        preserve_headers: If True and file exists, keep the existing header row

    Returns:
        Path to the reset Excel file
    """
    if not EXCEL_AVAILABLE or openpyxl is None:
        print("❌ Excel reset requires openpyxl library")
        return ""

    try:
        headers: List[str]
        if os.path.exists(master_excel_path) and preserve_headers:
            wb = openpyxl.load_workbook(master_excel_path)
            ws = wb.active
            if ws is None:
                raise RuntimeError("Worksheet not available")
            # Extract existing headers
            headers = [c.value for c in ws[1] if c.value]
        else:
            # Base headers (no dynamic rule columns)
            headers = [
                "Run ID", "Analysis Date", "Reviewer Name", "Target Name", "Datasite Location",
                "Document Name", "Knowledge Gaps", "Contract Start Date", "Type of Agreement",
                "Company 1", "Company 2",
                "Source Code Y/N", "Source Code Details", "Exclusivity Y/N", "Exclusivity Details",
                "Pricing Y/N", "Pricing Details", "IP Y/N", "IP Details",
                "Liability Y/N", "Liability Details", "Assignment Y/N", "Assignment Details", 
                "Secondary Use Rights", "Other Terms", "Comments"
            ]
            wb = openpyxl.Workbook()
            ws = wb.active

        # Rewrite headers and clear all data rows
        ws.title = ws.title or "Contract Analysis Comparison"
        # Clear sheet content
        ws.delete_rows(1, ws.max_row)
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Freeze header
        ws.freeze_panes = 'A2'

        # Save workbook
        os.makedirs(os.path.dirname(master_excel_path), exist_ok=True)
        wb.save(master_excel_path)
        print(f"🧹 Master template Excel reset: {master_excel_path}")
        return master_excel_path

    except Exception as e:
        print(f"❌ Error resetting master template Excel: {e}")
        return ""