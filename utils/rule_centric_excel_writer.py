#!/usr/bin/env python3
"""
Rule-Centric Excel Writer

Creates Excel output focused on rule compliance status, making it clear to reviewers
which rules are violated and why.
"""

import os
import json
from datetime import datetime
from typing import Dict, Any, List

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    Workbook = None
    Worksheet = None


def create_rule_compliance_spreadsheet(
    analysis_data: Dict[str, Any],
    output_path: str,
    target_document_path: str,
    rules_path: str
) -> str:
    """
    Create a rule-centric Excel spreadsheet showing compliance status for all rules.
    
    Args:
        analysis_data: Complete analysis results including rule_compliance_results
        output_path: Path where Excel file should be saved
        target_document_path: Path to analyzed document
        rules_path: Path to rules file
        
    Returns:
        Path to created Excel file
    """
    if not EXCEL_AVAILABLE or openpyxl is None:
        print("❌ Excel creation requires openpyxl library")
        return ""
    
    try:
        # Load all rules to ensure we include everything
        all_rules = _load_all_rules(rules_path)
        
        # Get compliance results from analysis
        compliance_results = analysis_data.get('rule_compliance_results', [])
        compliance_summary = analysis_data.get('rule_compliance_summary', {})
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create main compliance sheet
        _create_compliance_sheet(wb, all_rules, compliance_results, 
                               target_document_path, compliance_summary)
        
        # Create detailed findings sheet
        _create_findings_sheet(wb, compliance_results)
        
        # Create summary dashboard
        _create_summary_sheet(wb, compliance_summary, target_document_path)
        
        # Save workbook
        excel_path = output_path.replace('.yaml', '_compliance.xlsx').replace('.yml', '_compliance.xlsx')
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        wb.save(excel_path)
        
        print(f"✅ Rule compliance Excel created: {excel_path}")
        return excel_path
        
    except Exception as e:
        print(f"❌ Error creating compliance Excel: {e}")
        return ""


def _load_all_rules(rules_path: str) -> List[Dict[str, Any]]:
    """Load all rules from file"""
    if not os.path.exists(rules_path):
        return []
    
    with open(rules_path, 'r') as f:
        if rules_path.endswith('.json'):
            data = json.load(f)
        elif rules_path.endswith(('.yaml', '.yml')):
            import yaml
            data = yaml.safe_load(f)
        else:
            return []
    
    return data.get('rules', [])


def _create_compliance_sheet(wb: Workbook, 
                            all_rules: List[Dict[str, Any]],
                            compliance_results: List[Dict[str, Any]],
                            target_document_path: str,
                            summary: Dict[str, Any]) -> None:
    """Create the main compliance status sheet"""
    
    ws = wb.active
    ws.title = "Rule Compliance"
    
    # Create header rows
    headers = [
        "Rule ID",
        "Rule Name", 
        "Severity",
        "Compliance Status",
        "Confidence",
        "Rationale",
        "Evidence/Citations",
        "Exceptions Applied",
        "Action Required"
    ]
    
    # Document info row
    doc_name = os.path.basename(target_document_path)
    analysis_date = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    # Add document header
    ws.merge_cells('A1:I1')
    cell = ws['A1']
    cell.value = f"Rule Compliance Analysis: {doc_name}"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')
    
    # Add summary row
    ws.merge_cells('A2:I2')
    cell = ws['A2']
    score = summary.get('compliance_score', 0) * 100
    cell.value = (f"Overall Compliance Score: {score:.1f}% | "
                 f"Compliant: {summary.get('compliant', 0)} | "
                 f"Non-Compliant: {summary.get('non_compliant', 0)} | "
                 f"Review Required: {summary.get('requires_review', 0)}")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal='center')
    
    # Add analysis date
    ws.merge_cells('A3:I3')
    cell = ws['A3']
    cell.value = f"Analysis Date: {analysis_date}"
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal='center')
    
    # Add column headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Create lookup for compliance results
    results_by_id = {r['rule_id']: r for r in compliance_results}
    
    # Add data rows for ALL rules
    current_row = 5
    for rule in all_rules:
        rule_id = rule.get('id', 'unknown')
        rule_name = rule.get('name', rule_id)
        severity = rule.get('severity', 'medium')
        
        # Get compliance result if evaluated
        result = results_by_id.get(rule_id, {})
        
        if result:
            status = result.get('status', 'not_evaluated')
            confidence = result.get('confidence', 0)
            rationale = result.get('rationale', '')
            citations = result.get('citations', [])
            exceptions = result.get('exceptions_applied', [])
        else:
            # Rule was not evaluated
            status = 'not_evaluated'
            confidence = 0
            rationale = 'Rule was not evaluated'
            citations = []
            exceptions = []
        
        # Determine action required
        action = _determine_action(status, severity)
        
        # Format citations
        citation_text = ''
        if citations:
            citation_text = '\n'.join([c.get('quote', '') for c in citations[:3]])
        
        # Write row data
        row_data = [
            rule_id,
            rule_name,
            severity.upper(),
            status.replace('_', ' ').title(),
            f"{confidence:.0%}" if confidence > 0 else "N/A",
            rationale[:500] + '...' if len(rationale) > 500 else rationale,
            citation_text[:300] + '...' if len(citation_text) > 300 else citation_text,
            ', '.join(exceptions) if exceptions else 'None',
            action
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = value
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply conditional formatting
            if col_num == 4:  # Status column
                cell.fill = _get_status_fill(status)
                cell.font = Font(bold=True)
            elif col_num == 3:  # Severity column
                if severity == 'critical':
                    cell.font = Font(color='FF0000', bold=True)
                elif severity == 'high':
                    cell.font = Font(color='FF6600', bold=True)
            elif col_num == 9:  # Action column
                if 'Immediate' in action:
                    cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                    cell.font = Font(color='FF0000', bold=True)
                elif 'Review' in action:
                    cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        current_row += 1
    
    # Auto-adjust column widths
    column_widths = {
        'A': 15,  # Rule ID
        'B': 30,  # Rule Name
        'C': 10,  # Severity
        'D': 20,  # Status
        'E': 12,  # Confidence
        'F': 60,  # Rationale
        'G': 40,  # Evidence
        'H': 25,  # Exceptions
        'I': 25   # Action
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze panes below headers
    ws.freeze_panes = 'A5'


def _create_findings_sheet(wb: Workbook, compliance_results: List[Dict[str, Any]]) -> None:
    """Create detailed findings sheet for non-compliant rules"""
    
    ws = wb.create_sheet("Non-Compliance Details")
    
    # Filter for non-compliant and review-required rules
    issues = [r for r in compliance_results 
             if r.get('status') in ['non_compliant', 'partially_compliant', 'requires_review']]
    
    # Sort by severity
    severity_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
    issues.sort(key=lambda x: (severity_order.get(x.get('severity', 'medium'), 2), x.get('rule_name', '')))
    
    # Headers
    headers = [
        "Rule Name",
        "Severity", 
        "Status",
        "Detailed Findings",
        "Relevant Document Sections",
        "Recommended Actions"
    ]
    
    # Add title
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = f"Non-Compliance Findings ({len(issues)} issues identified)"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add findings
    current_row = 3
    for issue in issues:
        # Format relevant sections
        sections_text = ''
        if issue.get('relevant_sections'):
            sections_text = '\n\n'.join(issue['relevant_sections'][:3])
        
        # Determine recommended actions
        actions = _get_recommended_actions(issue)
        
        row_data = [
            issue.get('rule_name', issue.get('rule_id', 'Unknown')),
            issue.get('severity', 'medium').upper(),
            issue.get('status', '').replace('_', ' ').title(),
            issue.get('rationale', 'No details available'),
            sections_text,
            actions
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = value
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Highlight critical issues
            if col_num == 2 and value == 'CRITICAL':
                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
                cell.font = Font(color='FF0000', bold=True)
        
        current_row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 40
    
    ws.freeze_panes = 'A3'


def _create_summary_sheet(wb: Workbook, summary: Dict[str, Any], target_document_path: str) -> None:
    """Create executive summary dashboard"""
    
    ws = wb.create_sheet("Executive Summary")
    
    # Title
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "Compliance Analysis Executive Summary"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center')
    
    # Document info
    ws['A3'] = "Document:"
    ws['B3'] = os.path.basename(target_document_path)
    ws['A4'] = "Analysis Date:"
    ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    # Overall metrics
    ws['A6'] = "OVERALL METRICS"
    ws['A6'].font = Font(bold=True, size=12)
    
    metrics = [
        ("Total Rules Evaluated:", summary.get('total_rules', 0)),
        ("Overall Compliance Score:", f"{summary.get('compliance_score', 0)*100:.1f}%"),
        ("", ""),
        ("Compliant:", summary.get('compliant', 0)),
        ("Non-Compliant:", summary.get('non_compliant', 0)),
        ("Partially Compliant:", summary.get('partially_compliant', 0)),
        ("Not Applicable:", summary.get('not_applicable', 0)),
        ("Requires Review:", summary.get('requires_review', 0)),
    ]
    
    current_row = 7
    for label, value in metrics:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=2, value=value)
        
        # Format non-compliant row
        if label == "Non-Compliant:" and value > 0:
            ws.cell(row=current_row, column=2).font = Font(color='FF0000', bold=True)
        elif label == "Overall Compliance Score:":
            score_val = float(value.rstrip('%'))
            if score_val < 70:
                ws.cell(row=current_row, column=2).font = Font(color='FF0000', bold=True)
            elif score_val < 85:
                ws.cell(row=current_row, column=2).font = Font(color='FF6600', bold=True)
            else:
                ws.cell(row=current_row, column=2).font = Font(color='008000', bold=True)
        
        current_row += 1
    
    # Risk summary
    current_row += 1
    ws.cell(row=current_row, column=1, value="RISK SUMMARY")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 1
    
    risk_metrics = [
        ("Critical Issues:", summary.get('critical_issues', 0)),
        ("High Priority Issues:", summary.get('high_priority_issues', 0)),
    ]
    
    for label, value in risk_metrics:
        ws.cell(row=current_row, column=1, value=label)
        cell = ws.cell(row=current_row, column=2, value=value)
        if value > 0:
            cell.font = Font(color='FF0000' if 'Critical' in label else 'FF6600', bold=True)
        current_row += 1
    
    # Critical issues details
    if summary.get('critical_issue_details'):
        current_row += 1
        ws.cell(row=current_row, column=1, value="CRITICAL ISSUES REQUIRING IMMEDIATE ATTENTION")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color='FF0000')
        current_row += 1
        
        for issue in summary['critical_issue_details'][:5]:
            current_row += 1
            ws.cell(row=current_row, column=1, value=f"• {issue['rule_name']}:")
            ws.cell(row=current_row, column=2, value=issue['rationale'][:200])
            ws.merge_cells(f'B{current_row}:D{current_row}')
    
    # Recommendations
    current_row += 2
    ws.cell(row=current_row, column=1, value="RECOMMENDATIONS")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 1
    
    score = summary.get('compliance_score', 0) * 100
    if score < 60:
        recommendation = "⚠️ CRITICAL: Document has significant compliance issues. Legal review strongly recommended before proceeding."
        rec_color = 'FF0000'
    elif score < 80:
        recommendation = "⚠️ CAUTION: Document has notable compliance gaps. Review and negotiate amendments for high-priority issues."
        rec_color = 'FF6600'
    elif score < 95:
        recommendation = "✓ ACCEPTABLE: Document is largely compliant with minor issues. Address identified gaps if possible."
        rec_color = '008000'
    else:
        recommendation = "✓ EXCELLENT: Document shows strong compliance. Proceed with standard review process."
        rec_color = '008000'
    
    ws.cell(row=current_row, column=1, value=recommendation)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws.cell(row=current_row, column=1).font = Font(color=rec_color, bold=True)
    ws.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30


def _get_status_fill(status: str) -> PatternFill:
    """Get fill color based on compliance status"""
    fills = {
        'compliant': PatternFill(start_color='E6F3E6', end_color='E6F3E6', fill_type='solid'),  # Light green
        'non_compliant': PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'),  # Light red
        'partially_compliant': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),  # Light yellow
        'not_applicable': PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid'),  # Light gray
        'requires_review': PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid'),  # Light blue
        'not_evaluated': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White
    }
    return fills.get(status, fills['not_evaluated'])


def _determine_action(status: str, severity: str) -> str:
    """Determine required action based on status and severity"""
    if status == 'non_compliant':
        if severity == 'critical':
            return "Immediate legal review required"
        elif severity == 'high':
            return "Review and negotiate amendments"
        else:
            return "Document and assess impact"
    elif status == 'partially_compliant':
        if severity in ['critical', 'high']:
            return "Review gaps and negotiate"
        else:
            return "Monitor and document"
    elif status == 'requires_review':
        return "Manual review needed"
    elif status == 'not_evaluated':
        return "Pending evaluation"
    else:
        return "No action required"


def _get_recommended_actions(issue: Dict[str, Any]) -> str:
    """Get recommended actions for a non-compliance issue"""
    severity = issue.get('severity', 'medium')
    status = issue.get('status', '')
    rule_name = issue.get('rule_name', '')
    
    actions = []
    
    if severity == 'critical':
        actions.append("1. Escalate to legal counsel immediately")
        actions.append("2. Do not proceed without addressing this issue")
    elif severity == 'high':
        actions.append("1. Flag for negotiation priority")
        actions.append("2. Seek alternative language or amendments")
    
    if status == 'non_compliant':
        actions.append(f"3. Request specific amendment to address {rule_name}")
    elif status == 'requires_review':
        actions.append("3. Obtain expert opinion on interpretation")
    
    if not actions:
        actions.append("Document finding and assess business impact")
    
    return '\n'.join(actions)