#!/usr/bin/env python3
"""
Output Organization Utility

Manages structured output organization with:
- Run-specific directories for individual analysis
- Centralized Excel file for cross-document comparison
- Clean directory structure for easy navigation
"""

import os
import json
import pandas as pd
from datetime import datetime
from typing import Dict, Any, Tuple
from pathlib import Path
from utils.template_excel_writer import (
    update_master_template_excel
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class OutputOrganizer:
    """Manages structured output organization for contract analysis."""
    
    def __init__(self, base_output_dir: str = "data/output"):
        # Anchor relative paths to the project directory (agentic-process)
        if not os.path.isabs(base_output_dir):
            project_root = Path(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            self.base_dir = project_root / base_output_dir
        else:
            self.base_dir = Path(base_output_dir)
        self.runs_dir = self.base_dir / "runs"
        self.comparisons_dir = self.base_dir / "comparisons"
        self.archives_dir = self.base_dir / "archives"
        
        # Ensure directories exist
        for directory in [self.runs_dir, self.comparisons_dir, self.archives_dir]:
            directory.mkdir(parents=True, exist_ok=True)
    
    def create_run_directory(self, run_id: str, document_path: str, reference_path: str) -> Tuple[str, str]:
        """
        Create a run-specific directory and return paths for outputs.
        
        Args:
            run_id: Unique run identifier (e.g., "20250623_172309")
            document_path: Path to the target document
            reference_path: Path to the reference document
            
        Returns:
            Tuple of (run_directory_path, yaml_output_path)
        """
        run_dir = self.runs_dir / f"run_{run_id}"
        run_dir.mkdir(exist_ok=True)
        
        # Create metadata file
        metadata = {
            "run_id": run_id,
            "created_at": datetime.now().isoformat(),
            "document_info": {
                "target_document_path": str(document_path),
                "target_document_name": os.path.basename(document_path),
                "reference_document_path": str(reference_path),
                "reference_document_name": os.path.basename(reference_path)
            },
            "output_files": {
                "yaml": "analysis.yaml",
                "markdown": "analysis.md",
                "metadata": "metadata.json"
            }
        }
        
        metadata_path = run_dir / "metadata.json"
        with open(metadata_path, 'w') as f:
            json.dump(metadata, f, indent=2)
        
        yaml_output_path = run_dir / "analysis.yaml"
        
        print(f"📁 Created run directory: {run_dir}")
        return str(run_dir), str(yaml_output_path)
    
    def get_master_excel_path(self, use_template_format: bool = False) -> str:
        """Get the path to the master comparison Excel file."""
        if use_template_format:
            return str(self.comparisons_dir / "contract_analysis_template_master.xlsx")
        else:
            return str(self.comparisons_dir / "contract_analysis_master.xlsx")
    
    def get_document_identifier(self, document_path: str) -> str:
        """
        Generate a consistent identifier for a document.
        
        Args:
            document_path: Path to the document
            
        Returns:
            Consistent document identifier
        """
        # Use the filename without extension as the primary identifier
        filename = Path(document_path).stem
        
        # Clean up common problematic characters for Excel
        identifier = filename.replace(" ", "_").replace("-", "_")
        
        # Limit length to avoid Excel issues
        if len(identifier) > 50:
            identifier = identifier[:47] + "..."
        
        return identifier
    
    def update_master_excel(self, analysis_data: Dict[str, Any], run_id: str, 
                          document_path: str, reference_path: str, 
                          use_template_format: bool = False) -> str:
        """
        Update the master Excel file with new analysis data.
        Either adds a new row or updates existing row for the same document.
        
        Args:
            analysis_data: Complete analysis results
            run_id: Run identifier
            document_path: Path to analyzed document
            reference_path: Path to reference document
            use_template_format: Whether to use template format
            
        Returns:
            Path to the updated Excel file
        """
        if use_template_format:
            # Use the specialized template Excel writer (includes dynamic rule columns)
            excel_path = self.get_master_excel_path(use_template_format=True)
            result = update_master_template_excel(
                analysis_data, run_id, document_path, reference_path, excel_path
            )
            # Also update Due Diligence two-row header sheet
            try:
                import openpyxl
                from utils.template_excel_writer import _update_due_diligence_sheet
                wb = openpyxl.load_workbook(excel_path)
                _update_due_diligence_sheet(wb, analysis_data, run_id, document_path)
                wb.save(excel_path)
            except Exception:
                pass
            if not result:
                print("⚠️ Master template update returned empty path. Attempting to create a new master file.")
                # Try to reset and write headers, then retry
                try:
                    from utils.template_excel_writer import reset_master_template_excel
                    reset_master_template_excel(excel_path, preserve_headers=False)
                    result = update_master_template_excel(
                        analysis_data, run_id, document_path, reference_path, excel_path
                    )
                except Exception as e:
                    print(f"❌ Could not reset master template: {e}")
            return result
        
        # Use the existing generic format
        excel_path = self.get_master_excel_path(use_template_format=False)
        document_id = self.get_document_identifier(document_path)
        
        # Extract key data from analysis
        exec_summary = analysis_data.get('executive_summary', {})
        validation = analysis_data.get('validation_results', {}).get('validation_summary', {})
        risk_assessment = analysis_data.get('risk_assessment', {}).get('overall_risk_assessment', {})
        quality_metrics = analysis_data.get('quality_metrics', {})
        
        # Create new row data
        new_row = {
            'Document_ID': document_id,
            'Document_Name': os.path.basename(document_path),
            'Document_Path': document_path,
            'Reference_Document': os.path.basename(reference_path),
            'Run_ID': run_id,
            'Analysis_Date': exec_summary.get('analysis_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            'Risk_Level': exec_summary.get('risk_level', 'Unknown'),
            'Quality_Grade': quality_metrics.get('quality_grade', 'Unknown'),
            'Quality_Score': quality_metrics.get('overall_quality_score', 0),
            'Overall_Recommendation': exec_summary.get('overall_recommendation', 'Unknown'),
            'Total_Questions': validation.get('total_questions', 0),
            'Questions_Passed': validation.get('questions_passed', 0),
            'Questions_Failed': validation.get('questions_failed', 0),
            'Questions_With_Warnings': validation.get('questions_with_warnings', 0),
            'Manual_Review_Required': 'Yes' if quality_metrics.get('manual_review_required', False) else 'No',
            'Total_Red_Flags': analysis_data.get('risk_assessment', {}).get('red_flag_summary', {}).get('total_red_flags', 0),
            'Critical_Issues': analysis_data.get('risk_assessment', {}).get('red_flag_summary', {}).get('severity_breakdown', {}).get('Critical', 0),
            'High_Risk_Issues': analysis_data.get('risk_assessment', {}).get('red_flag_summary', {}).get('severity_breakdown', {}).get('High', 0),
            'Average_Risk_Score': risk_assessment.get('average_risk_score', 0),
            'Processing_Success_Rate': analysis_data.get('processing_metadata', {}).get('overall_classification_success_rate', '0%'),
            'Run_Directory': f"runs/run_{run_id}",
            'Last_Updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Load existing Excel file or create new DataFrame
        try:
            if os.path.exists(excel_path):
                df = pd.read_excel(excel_path)
                
                # Check if document already exists
                existing_mask = df['Document_ID'] == document_id
                if existing_mask.any():
                    # Update existing row
                    for col, value in new_row.items():
                        df.loc[existing_mask, col] = value
                    print(f"📊 Updated existing row for document: {document_id}")
                else:
                    # Add new row
                    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                    print(f"📊 Added new row for document: {document_id}")
            else:
                # Create new DataFrame
                df = pd.DataFrame([new_row])
                print(f"📊 Created new master Excel file with document: {document_id}")
            
            # Sort by most recent analysis date
            df = df.sort_values('Analysis_Date', ascending=False)
            
            # Write Excel file with formatting
            if EXCEL_AVAILABLE:
                self._write_formatted_excel(df, excel_path)
            else:
                # Fallback to CSV
                csv_path = excel_path.replace('.xlsx', '.csv')
                df.to_csv(csv_path, index=False)
                excel_path = csv_path
                print(f"⚠️ Excel not available, saved as CSV: {csv_path}")
            
            print(f"📊 Master comparison file updated: {excel_path}")
            print(f"📈 Total documents tracked: {len(df)}")
            
            return excel_path
            
        except Exception as e:
            print(f"❌ Error updating master Excel file: {e}")
            return ""
    
    def _write_formatted_excel(self, df: pd.DataFrame, excel_path: str) -> None:
        """Write DataFrame to Excel with professional formatting."""
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Contract Analysis Comparison')
            
            # Get the workbook and worksheet
            worksheet = writer.sheets['Contract Analysis Comparison']
            
            # Header formatting
            header_font = Font(color='FFFFFF', bold=True)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)  # Cap at 40 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Conditional formatting for risk levels
            high_risk_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')  # Light red
            medium_risk_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light yellow
            low_risk_fill = PatternFill(start_color='E6F3E6', end_color='E6F3E6', fill_type='solid')  # Light green
            
            # Apply conditional formatting
            # Find Risk_Level column index safely
            risk_level_col_idx = None
            for idx, col_name in enumerate(df.columns):
                if col_name == 'Risk_Level':
                    risk_level_col_idx = idx + 1  # Excel columns are 1-based
                    break
            
            if risk_level_col_idx:
                for row_num in range(2, len(df) + 2):  # Start from row 2 (after header)
                    risk_level_cell = worksheet.cell(row=row_num, column=risk_level_col_idx)
                    risk_level = str(risk_level_cell.value).upper()
                    
                    if risk_level in ['HIGH', 'CRITICAL']:
                        for col_num in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row_num, column=col_num).fill = high_risk_fill
                    elif risk_level == 'MEDIUM':
                        for col_num in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row_num, column=col_num).fill = medium_risk_fill
                    elif risk_level == 'LOW':
                        for col_num in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row_num, column=col_num).fill = low_risk_fill
            
            # Freeze the header row
            worksheet.freeze_panes = 'A2'
    
    def create_summary_dashboard(self) -> str:
        """
        Create a markdown summary dashboard of all analyses.
        
        Returns:
            Path to the created dashboard file
        """
        dashboard_path = self.comparisons_dir / "summary_dashboard.md"
        
        try:
            # Load master Excel data
            excel_path = self.get_master_excel_path()
            if not os.path.exists(excel_path):
                return ""
            
            df = pd.read_excel(excel_path)
            
            # Generate dashboard content
            dashboard_content = f"""# Contract Analysis Dashboard

*Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*

## Overview

**Total Documents Analyzed**: {len(df)}
**Latest Analysis**: {df['Analysis_Date'].max() if not df.empty else 'None'}

## Risk Distribution

"""
            
            # Risk level distribution
            if not df.empty:
                risk_counts = df['Risk_Level'].value_counts()
                for risk_level, count in risk_counts.items():
                    percentage = (count / len(df)) * 100
                    dashboard_content += f"- **{risk_level}**: {count} documents ({percentage:.1f}%)\n"
            
            dashboard_content += "\n## Quality Overview\n\n"
            
            # Quality distribution
            if not df.empty:
                quality_grades = df['Quality_Grade'].value_counts()
                avg_quality = df['Quality_Score'].mean()
                dashboard_content += f"**Average Quality Score**: {avg_quality:.1f}%\n\n"
                
                for grade, count in quality_grades.items():
                    percentage = (count / len(df)) * 100
                    dashboard_content += f"- **Grade {grade}**: {count} documents ({percentage:.1f}%)\n"
            
            dashboard_content += "\n## Recent Analyses\n\n"
            
            # Recent analyses (last 10)
            if not df.empty:
                recent_df = df.head(10)
                dashboard_content += "| Document | Risk Level | Quality | Analysis Date | Review Required |\n"
                dashboard_content += "|----------|------------|---------|---------------|------------------|\n"
                
                for _, row in recent_df.iterrows():
                    doc_name = row['Document_Name'][:30] + "..." if len(row['Document_Name']) > 30 else row['Document_Name']
                    dashboard_content += f"| {doc_name} | {row['Risk_Level']} | {row['Quality_Grade']} ({row['Quality_Score']:.0f}%) | {row['Analysis_Date']} | {row['Manual_Review_Required']} |\n"
            
            dashboard_content += "\n## Documents Requiring Review\n\n"
            
            # Documents requiring manual review
            if not df.empty:
                review_required = df[df['Manual_Review_Required'] == 'Yes']
                if not review_required.empty:
                    for _, row in review_required.iterrows():
                        dashboard_content += f"- **{row['Document_Name']}** (Risk: {row['Risk_Level']}, Quality: {row['Quality_Grade']})\n"
                else:
                    dashboard_content += "No documents currently require manual review.\n"
            
            dashboard_content += f"\n---\n*Generated automatically from {len(df)} contract analyses*\n"
            
            # Write dashboard file
            with open(dashboard_path, 'w', encoding='utf-8') as f:
                f.write(dashboard_content)
            
            print(f"📊 Summary dashboard created: {dashboard_path}")
            return str(dashboard_path)
            
        except Exception as e:
            print(f"❌ Error creating summary dashboard: {e}")
            return ""
    
    def archive_old_outputs(self) -> None:
        """Move old flat-structure outputs to archives directory."""
        archive_flat_dir = self.archives_dir / "old_flat_structure"
        archive_flat_dir.mkdir(exist_ok=True)
        
        # Look for files that match the old naming pattern
        old_pattern_files = list(self.base_dir.glob("*_analysis_*.yaml"))
        old_pattern_files.extend(list(self.base_dir.glob("*_analysis_*.md")))
        old_pattern_files.extend(list(self.base_dir.glob("*_analysis_*.xlsx")))
        
        moved_count = 0
        for file_path in old_pattern_files:
            if file_path.parent == self.base_dir:  # Only move files directly in base_dir
                try:
                    new_path = archive_flat_dir / file_path.name
                    file_path.rename(new_path)
                    moved_count += 1
                except Exception as e:
                    print(f"⚠️ Could not archive {file_path.name}: {e}")
        
        if moved_count > 0:
            print(f"📦 Archived {moved_count} old output files to: {archive_flat_dir}")

# Global instance
output_organizer = OutputOrganizer() 