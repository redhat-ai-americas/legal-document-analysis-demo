import os
import csv
from typing import Dict, Any, List

import yaml
import json

try:
	import openpyxl  # type: ignore
except Exception:
	openpyxl = None

from workflows.state import ContractAnalysisState
from utils.rules_schema import normalize_header, normalize_rule_row, validate_rules


def _read_csv(path: str) -> List[Dict[str, Any]]:
	with open(path, newline="", encoding="utf-8") as f:
		reader = csv.DictReader(f)
		reader.fieldnames = [normalize_header(h) for h in (reader.fieldnames or [])]
		raw = [{k: v for k, v in row.items()} for row in reader]
		return [normalize_rule_row(r) for r in raw]


def _read_yaml(path: str) -> List[Dict[str, Any]]:
	with open(path, "r", encoding="utf-8") as f:
		data = yaml.safe_load(f) or []
		if isinstance(data, dict) and "rules" in data:
			items = data["rules"]
		elif isinstance(data, list):
			items = data
		else:
			items = []
		return [normalize_rule_row({k: v for k, v in item.items()}) for item in items]


def _read_xlsx(path: str) -> List[Dict[str, Any]]:
	if openpyxl is None:
		raise RuntimeError("openpyxl not installed; cannot parse XLSX rules")
	wb = openpyxl.load_workbook(path)
	sheet = wb.active
	headers = [normalize_header(c.value or "") for c in next(sheet.iter_rows(min_row=1, max_row=1))]
	rules: List[Dict[str, Any]] = []
	for row in sheet.iter_rows(min_row=2, values_only=True):
		row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
		rules.append(normalize_rule_row(row_dict))
	return rules


def _read_json(path: str) -> List[Dict[str, Any]]:
	with open(path, "r", encoding="utf-8") as f:
		data = json.load(f)
		if isinstance(data, dict) and "rules" in data:
			items = data["rules"]
		elif isinstance(data, list):
			items = data
		else:
			items = []
		return [normalize_rule_row({k: v for k, v in item.items()}) for item in items]


def rules_loader(state: ContractAnalysisState) -> Dict[str, Any]:
	"""Load and normalize rules file into state as rules_data.

	Expects state["rules_path"] to be set to a CSV/XLSX/YAML/JSON file.
	"""
	rules_path = state.get("rules_path") or ""
	print(f"  🧭 rules_loader: incoming rules_path={rules_path}")
	if not rules_path:
		raise ValueError("rules_path must be provided in state")
	if not os.path.exists(rules_path):
		raise FileNotFoundError(f"Rules file not found: {rules_path}")

	ext = os.path.splitext(rules_path)[1].lower()
	if ext in (".csv",):
		items = _read_csv(rules_path)
	elif ext in (".yml", ".yaml"):
		items = _read_yaml(rules_path)
	elif ext in (".xlsx", ".xlsm"):
		items = _read_xlsx(rules_path)
	elif ext in (".json",):
		items = _read_json(rules_path)
	else:
		raise ValueError(f"Unsupported rules file type: {ext}")

	# Dedupe rules by (id, rule_text)
	seen = set()
	deduped: List[Dict[str, Any]] = []
	for r in items:
		key = (str(r.get("id", "")).strip(), str(r.get("rule_text", "")).strip())
		if key in seen:
			continue
		seen.add(key)
		deduped.append(r)

	valid, issues = validate_rules(deduped)
	if issues:
		# Keep processing but surface validation issues in state warnings
		print("Rules validation issues:")
		for issue in issues:
			print("  -", issue)

	if True:
		print(f"  📜 Loaded rules: {len(valid)} from {rules_path}")
		if len(valid) == 0:
			state_warnings = state.get("warnings", []) or []
			state_warnings.append(f"No rules loaded from {rules_path}; compliance will be unknown.")
			state["warnings"] = state_warnings

	return {"rules_data": valid, "rules_path": rules_path}
