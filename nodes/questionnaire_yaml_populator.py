import yaml
import os
import pandas as pd
from datetime import datetime
from typing import Dict, Any, List, Sequence
from workflows.state import ContractAnalysisState, ClassifiedSentence
from utils.quality_analyzer import quality_analyzer
from utils.risk_assessor import risk_assessor
from utils.red_flag_detector import red_flag_detector
from utils.citation_tracker import citation_tracker
from utils.output_organizer import output_organizer
from utils.template_excel_writer import create_template_excel

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available, Excel export will fall back to CSV")

def extract_deterministic_fields(state: ContractAnalysisState) -> Dict[str, str]:
    """
    Extract fields that can be determined programmatically without LLM processing.
    Now uses entity extraction results when available.
    """
    target_path = state.get('target_document_path', '')
    extracted_entities = state.get('extracted_entities', {})
    
    # Start with basic deterministic data
    deterministic_data = {
        'document_name': os.path.basename(target_path) if target_path else 'Unknown',
        'datasite_location': target_path,
        'reviewer_name': 'Automated Analysis',
        'processing_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    # Add extracted entity data with confidence information
    if extracted_entities:
        # Contract start date
        start_date = extracted_entities.get('contract_start_date')
        if start_date:
            deterministic_data['contract_start_date'] = f"{start_date['standardized_date']} (extracted: {start_date['extraction_method']}, confidence: {start_date['confidence']}%)"
        else:
            deterministic_data['contract_start_date'] = 'Not found through automatic extraction'
            
        # Target company name (Party 1)
        party1 = extracted_entities.get('party1')
        if party1:
            deterministic_data['target_company_name'] = f"{party1['name']} (extracted: {party1['extraction_method']}, confidence: {party1['confidence']}%)"
        else:
            deterministic_data['target_company_name'] = 'Not found through automatic extraction'
            
        # Counterparty name (Party 2)
        party2 = extracted_entities.get('party2')
        if party2:
            deterministic_data['counterparty_name'] = f"{party2['name']} (extracted: {party2['extraction_method']}, confidence: {party2['confidence']}%)"
        else:
            deterministic_data['counterparty_name'] = 'Not found through automatic extraction'
    
    return deterministic_data

def extract_questionnaire_data(questionnaire_responses: Dict[str, Any], deterministic_fields: Dict[str, str]) -> Dict[str, Any]:
    """
    Extract questionnaire responses and organize them in a structured format for YAML output.
    """
    output_data = {
        'document_information': {},
        'contract_summary': {},
        'gaps_and_comments': {},
        'processing_metadata': {
            'processing_start_time': None,
            'processing_end_time': None,
            'total_processing_time': None,
            'model_performance': {},
            'error_summary': {
                'total_errors': 0,
                'errors_by_model': {},
                'error_types': {}
            }
        },
        'risk_summary': {},
        'quality_metrics': {},
        'validation_results': {
            'questions_processed': 0,
            'questions_with_errors': 0,
            'questions_with_low_confidence': 0,
            'questions_requiring_review': []
        }
    }
    
    # Section mapping for better organization
    section_mapping = {
        'part_1_document_information': 'document_information',
        'part_2_contract_summary': 'contract_summary', 
        'part_3_gaps_internal_comments': 'gaps_and_comments'
    }
    
    # Extract answers from questionnaire responses
    for section_key, section_data in questionnaire_responses.items():
        yaml_section = section_mapping.get(section_key, 'other')
        
        for question in section_data.get('questions', []):
            question_id = question.get('id', '')
            answer = question.get('answer', 'Not specified')
            
            # Use deterministic value if available
            if question_id in deterministic_fields:
                answer = deterministic_fields[question_id]
            
            # Store with enhanced information including confidence, citations, and risk
            question_entry = {
                'question': question.get('prompt', ''),
                'answer': answer,
                'guideline': question.get('guideline', ''),
                'confidence': question.get('confidence', 0),
                'extraction_confidence': question.get('extraction_confidence', {}),
                'decision_attribution': question.get('decision_attribution', {}),
                'citations': [],
                'risk_assessment': None,
                'validation_status': {
                    'has_error': False,
                    'has_low_confidence': question.get('confidence', 0) < 0.7,
                    'requires_review': False,
                    'validation_messages': []
                },
                'model_analysis': {
                    'granite': question.get('model_responses', {}).get('granite', {}),
                    'comparison': question.get('model_responses', {}).get('comparison', {})
                }
            }
            
            # Add citation information if available
            citation_ids = question.get('citations', [])
            if citation_ids:
                question_entry['citations'] = citation_tracker.format_citations_for_yaml(citation_ids)
            
            # Add risk assessment if available
            risk_assessment = question.get('risk_assessment')
            if risk_assessment:
                question_entry['risk_assessment'] = {
                    'risk_level': risk_assessment.get('risk_level'),
                    'risk_score': risk_assessment.get('risk_score'),
                    'business_impact': risk_assessment.get('business_impact'),
                    'mitigation_urgency': risk_assessment.get('mitigation_urgency'),
                    'red_flags': risk_assessment.get('red_flags', []),
                    'recommendations': risk_assessment.get('recommendations', [])
                }
            
            # Get model responses if available
            model_responses = question.get('model_responses', {})
            granite_response = model_responses.get('granite', {})
            model_responses.get('comparison', {})
            
            # Track validation status
            has_error = False
            has_low_confidence = question.get('confidence', 0) < 0.7
            validation_messages = []
            
            # Check for errors in model responses
            for model_name, response in [('granite', granite_response)]:
                if 'error' in response:
                    has_error = True
                    error_msg = response.get('error', 'Unknown error')
                    validation_messages.append(f"{model_name} error: {error_msg}")
                    output_data['processing_metadata']['error_summary']['total_errors'] += 1
                    output_data['processing_metadata']['error_summary']['errors_by_model'][model_name] = \
                        output_data['processing_metadata']['error_summary']['errors_by_model'].get(model_name, 0) + 1
            
            # Check for low confidence
            if has_low_confidence:
                validation_messages.append(f"Low confidence score ({question.get('confidence', 0):.2f})")
                output_data['validation_results']['questions_with_low_confidence'] += 1
            
            # Update validation status
            question_entry['validation_status'].update({
                'has_error': has_error,
                'has_low_confidence': has_low_confidence,
                'requires_review': has_error or has_low_confidence,
                'validation_messages': validation_messages
            })
            
            # Update validation tracking
            output_data['validation_results']['questions_processed'] += 1
            if has_error:
                output_data['validation_results']['questions_with_errors'] += 1
            
            if has_error or has_low_confidence:
                output_data['validation_results']['questions_requiring_review'].append({
                    'question_id': question_id,
                    'section': yaml_section,
                    'reason': ', '.join(validation_messages)
                })
            
            output_data[yaml_section][question_id] = question_entry
    
    return output_data

def create_executive_summary(state: ContractAnalysisState, quality_report: Dict[str, Any], 
                           risk_profile: Dict[str, Any], red_flag_analysis: Dict[str, Any]) -> Dict[str, Any]:
    """
    Create an executive summary of the contract analysis.
    """
    # Extract key findings
    overall_risk = red_flag_analysis.get("overall_risk_level", "UNKNOWN")
    quality_grade = quality_report.get("quality_grade", "F")
    quality_score = quality_report.get("overall_quality_score", 0)
    
    # Key contract details
    target_document = state.get('target_document_path', 'Unknown')
    document_name = os.path.basename(target_document) if target_document else 'Unknown'
    
    # Risk summary
    critical_red_flags = red_flag_analysis.get("severity_breakdown", {}).get("Critical", 0)
    high_red_flags = red_flag_analysis.get("severity_breakdown", {}).get("High", 0)
    
    # Generate recommendation
    if critical_red_flags > 0:
        recommendation = "DO NOT SIGN - Critical issues require immediate resolution"
    elif high_red_flags >= 3:
        recommendation = "HIGH CAUTION - Comprehensive legal review required before signing"
    elif overall_risk in ["HIGH", "CRITICAL"]:
        recommendation = "CAUTION - Address risk provisions before signing"
    elif quality_score < 60:
        recommendation = "UNCERTAIN - Low analysis confidence, manual review required"
    else:
        recommendation = "ACCEPTABLE - Standard legal review recommended"
    
    # Key attention items
    attention_items = []
    if critical_red_flags > 0:
        attention_items.append(f"{critical_red_flags} critical red flag(s) detected")
    if high_red_flags > 0:
        attention_items.append(f"{high_red_flags} high-risk issue(s) identified")
    if quality_score < 70:
        attention_items.append(f"Low analysis confidence ({quality_score:.0f}%)")
    
    # High-risk areas
    high_risk_areas = []
    risk_categories = red_flag_analysis.get("category_breakdown", {})
    for category, count in risk_categories.items():
        if count > 0:
            category_name = category.replace("_", " ").title()
            high_risk_areas.append(f"{category_name} ({count} issue{'s' if count > 1 else ''})")
    
    return {
        "document_name": document_name,
        "analysis_date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "overall_recommendation": recommendation,
        "risk_level": overall_risk,
        "quality_grade": f"{quality_grade} ({quality_score:.0f}%)",
        "requires_attention": attention_items,
        "high_risk_areas": high_risk_areas[:5],  # Top 5
        "key_statistics": {
            "total_red_flags": red_flag_analysis.get("total_red_flags", 0),
            "critical_issues": critical_red_flags,
            "high_risk_issues": high_red_flags,
            "analysis_confidence": f"{quality_score:.0f}%"
        },
        "next_steps": _generate_next_steps(overall_risk, critical_red_flags, high_red_flags, quality_score)
    }

def _generate_next_steps(risk_level: str, critical_flags: int, high_flags: int, quality_score: float) -> List[str]:
    """Generate recommended next steps based on analysis results."""
    steps = []
    
    if critical_flags > 0:
        steps.append("1. IMMEDIATE: Review critical red flags before any further action")
        steps.append("2. URGENT: Consult legal counsel regarding critical issues")
        steps.append("3. Consider rejecting contract or demanding major revisions")
    elif high_flags >= 3:
        steps.append("1. Schedule comprehensive legal review")
        steps.append("2. Prepare negotiation strategy for high-risk provisions")
        steps.append("3. Consider alternative contract terms")
    elif risk_level in ["HIGH", "CRITICAL"]:
        steps.append("1. Review highlighted risk provisions")
        steps.append("2. Negotiate key terms before acceptance")
    else:
        steps.append("1. Conduct standard legal review")
        steps.append("2. Verify key terms and conditions")
    
    if quality_score < 70:
        steps.append("4. Manual review recommended due to low analysis confidence")
    
    steps.append("5. Document any negotiated changes")
    steps.append("6. Obtain final legal approval before signing")
    
    return steps

def _generate_validation_checklist(quality_report: Dict[str, Any], risk_summary: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Generate a validation checklist for human reviewers."""
    checklist = []
    
    # Quality-based checks
    overall_score = quality_report.get("overall_quality_score", 0)
    if overall_score < 70:
        checklist.append({
            "priority": "High",
            "item": "Verify analysis accuracy due to low confidence score",
            "description": f"Overall quality score: {overall_score:.0f}%"
        })
    
    # Risk-based checks
    red_flags = risk_summary.get("red_flag_summary", {}).get("total_red_flags", 0)
    if red_flags > 0:
        checklist.append({
            "priority": "Critical" if red_flags >= 3 else "High",
            "item": f"Review {red_flags} identified red flag(s)",
            "description": "Manual verification of automated risk detection"
        })
    
    # Standard checks
    checklist.extend([
        {
            "priority": "Medium",
            "item": "Verify key dates and party information",
            "description": "Confirm extracted entity information is accurate"
        },
        {
            "priority": "Medium", 
            "item": "Review liability and indemnification provisions",
            "description": "Manual review of critical legal terms"
        },
        {
            "priority": "Low",
            "item": "Check for missing or incomplete sections",
            "description": "Ensure all contract sections were analyzed"
        }
    ])
    
    return checklist

def _generate_reviewer_guidance(quality_report: Dict[str, Any], risk_summary: Dict[str, Any]) -> List[str]:
    """Generate guidance for human reviewers."""
    guidance = []
    
    # Quality-based guidance
    overall_score = quality_report.get("overall_quality_score", 0)
    if overall_score < 60:
        guidance.append("LOW CONFIDENCE: This analysis has low confidence scores. Manual review of all findings is strongly recommended.")
    elif overall_score < 80:
        guidance.append("MEDIUM CONFIDENCE: Review highlighted areas and verify key findings manually.")
    
    # Risk-based guidance
    risk_level = risk_summary.get("overall_risk_assessment", {}).get("document_risk_level", "LOW")
    if risk_level in ["CRITICAL", "HIGH"]:
        guidance.append(f"{risk_level} RISK: This contract contains significant risk factors requiring legal review.")
    
    # Red flag guidance
    critical_flags = risk_summary.get("red_flag_summary", {}).get("severity_breakdown", {}).get("Critical", 0)
    if critical_flags > 0:
        guidance.append(f"CRITICAL RED FLAGS: {critical_flags} critical issue(s) detected. Immediate legal consultation recommended.")
    
    # Citation guidance
    guidance.append("CITATIONS: Use provided citations to verify extracted information against source text.")
    guidance.append("CONFIDENCE SCORES: Lower confidence scores indicate areas requiring additional scrutiny.")
    
    return guidance

def create_risk_summary(state: ContractAnalysisState) -> Dict[str, Any]:
    """
    Create comprehensive risk summary for the document.
    """
    classified_sentences: Sequence[ClassifiedSentence] = state.get('classified_sentences', [])
    
    # Convert ClassifiedSentence to Dict[str, Any] for utility functions
    dict_sentences = [dict(sentence) for sentence in classified_sentences]
    
    # Get overall risk profile
    risk_profile = risk_assessor.assess_document_risk_profile(dict_sentences)
    
    # Get red flag analysis
    red_flag_analysis = red_flag_detector.analyze_document_red_flags(dict_sentences)
    
    # Combine for comprehensive risk summary
    return {
        "overall_risk_assessment": {
            "document_risk_level": risk_profile.get("document_risk_level"),
            "average_risk_score": risk_profile.get("average_risk_score"),
            "maximum_risk_score": risk_profile.get("maximum_risk_score"),
            "high_risk_clause_count": risk_profile.get("high_risk_clause_count"),
            "recommendation": red_flag_analysis.get("overall_recommendation")
        },
        "red_flag_summary": {
            "total_red_flags": red_flag_analysis.get("total_red_flags", 0),
            "severity_breakdown": red_flag_analysis.get("severity_breakdown", {}),
            "category_breakdown": red_flag_analysis.get("category_breakdown", {}),
            "priority_flags": red_flag_analysis.get("priority_flags", [])[:5]  # Top 5
        },
        "risk_categories": risk_profile.get("risk_categories", {}),
        "business_impact_summary": red_flag_analysis.get("business_impact_summary", {}),
        "action_items": red_flag_analysis.get("action_items", [])[:8],  # Top 8
        "negotiation_priorities": red_flag_analysis.get("negotiation_priorities", [])
    }

def calculate_processing_stats(state: ContractAnalysisState) -> Dict[str, Any]:
    """
    Calculate processing statistics for the analysis.
    """
    classified_sentences = state.get('classified_sentences', [])
    reference_classified = state.get('reference_classified_sentences', [])
    
    # Count successful classifications
    target_successful = len([s for s in classified_sentences if s.get('classes')])
    total_target = len(classified_sentences)
    
    reference_successful = len([s for s in reference_classified if s.get('classes')])
    total_reference = len(reference_classified)
    
    # Calculate success rate
    total_sentences = total_target + total_reference
    total_successful = target_successful + reference_successful
    success_rate = (total_successful / total_sentences * 100) if total_sentences > 0 else 0
    
    return {
        'target_sentences_processed': total_target,
        'target_sentences_classified': target_successful,
        'reference_sentences_processed': total_reference,
        'reference_sentences_classified': reference_successful,
        'overall_classification_success_rate': f"{success_rate:.1f}%",
        'processing_timestamp': datetime.now().isoformat()
    }

def write_markdown_report(analysis_data: Dict[str, Any], output_path: str, target_document_path: str, reference_document_path: str) -> str:
    """
    Write a well-structured markdown report for human review.
    
    Args:
        analysis_data: The complete analysis data
        output_path: Base path for the YAML file (will change extension to .md)
        target_document_path: Path to analyzed document
        reference_document_path: Path to reference document
        
    Returns:
        Path to the created markdown file
    """
    # Generate markdown file path
    md_path = output_path.replace('.yaml', '.md').replace('.yml', '.md')
    
    # Extract key data
    exec_summary = analysis_data.get('executive_summary', {})
    doc_info = analysis_data.get('document_information', {})
    contract_summary = analysis_data.get('contract_summary', {})
    gaps = analysis_data.get('gaps_and_comments', {})
    validation = analysis_data.get('validation_results', {})
    risk_assessment = analysis_data.get('risk_assessment', {})
    quality_metrics = analysis_data.get('quality_metrics', {})
    
    # Create markdown content
    md_content = f"""# Contract Analysis Report

## Document Information
- **Document**: {exec_summary.get('document_name', 'Unknown')}
- **Analysis Date**: {exec_summary.get('analysis_date', 'Unknown')}
- **Risk Level**: {exec_summary.get('risk_level', 'Unknown')}
- **Quality Grade**: {exec_summary.get('quality_grade', 'Unknown')}
- **Overall Recommendation**: {exec_summary.get('overall_recommendation', 'Unknown')}

## Executive Summary

### Key Statistics
"""
    
    # Add key statistics
    key_stats = exec_summary.get('key_statistics', {})
    for stat, value in key_stats.items():
        stat_name = stat.replace('_', ' ').title()
        md_content += f"- **{stat_name}**: {value}\n"
    
    # Add attention items
    attention_items = exec_summary.get('requires_attention', [])
    if attention_items:
        md_content += "\n### Requires Attention\n"
        for item in attention_items:
            md_content += f"- {item}\n"
    
    # Add high-risk areas
    high_risk_areas = exec_summary.get('high_risk_areas', [])
    if high_risk_areas:
        md_content += "\n### High-Risk Areas\n"
        for area in high_risk_areas:
            md_content += f"- {area}\n"
    
    # Add next steps
    next_steps = exec_summary.get('next_steps', [])
    if next_steps:
        md_content += "\n### Recommended Next Steps\n"
        for step in next_steps:
            md_content += f"{step}\n"
    
    # Document Information Section
    md_content += "\n## Document Details\n\n"
    for field_id, field_data in doc_info.items():
        if isinstance(field_data, dict):
            question = field_data.get('question', field_id)
            answer = field_data.get('answer', 'Not specified')
            confidence = field_data.get('confidence', 0)
            
            md_content += f"### {question}\n"
            md_content += f"**Answer**: {answer}\n"
            if confidence > 0.8:
                md_content += f"*Confidence: High ({confidence:.0%})*\n\n"
            elif confidence > 0.6:
                md_content += f"*Confidence: Medium ({confidence:.0%})*\n\n"
            else:
                md_content += f"*Confidence: Low ({confidence:.0%})*\n\n"
            md_content += "\n"
    
    # Contract Summary
    md_content += "\n## Contract Summary\n\n"
    for item_id, item_data in contract_summary.items():
        if isinstance(item_data, dict):
            question = item_data.get('question', item_id)
            answer = item_data.get('answer', 'Not specified')
            confidence = item_data.get('confidence', 0)
            
            md_content += f"### {question}\n"
            md_content += f"{answer}\n"
            if confidence > 0.8:
                md_content += f"*Confidence: High ({confidence:.0%})*\n\n"
            elif confidence > 0.6:
                md_content += f"*Confidence: Medium ({confidence:.0%})*\n\n"
            else:
                md_content += f"*Confidence: Low ({confidence:.0%})*\n\n"
            md_content += "\n"
    
    # Risk Assessment
    overall_risk = risk_assessment.get('overall_risk_assessment', {})
    red_flag_summary = risk_assessment.get('red_flag_summary', {})
    
    md_content += "\n## Risk Assessment\n\n"
    md_content += f"- **Document Risk Level**: {overall_risk.get('document_risk_level', 'Unknown')}\n"
    md_content += f"- **Average Risk Score**: {overall_risk.get('average_risk_score', 0)}\n"
    md_content += f"- **Total Red Flags**: {red_flag_summary.get('total_red_flags', 0)}\n"
    
    # Red flag breakdown
    severity_breakdown = red_flag_summary.get('severity_breakdown', {})
    if any(severity_breakdown.values()):
        md_content += "\n### Red Flag Breakdown\n"
        for severity, count in severity_breakdown.items():
            if count > 0:
                md_content += f"- **{severity}**: {count}\n"
    
    # Quality Metrics
    md_content += "\n## Quality Metrics\n\n"
    md_content += f"- **Overall Quality Score**: {quality_metrics.get('overall_quality_score', 0):.1f}%\n"
    md_content += f"- **Quality Grade**: {quality_metrics.get('quality_grade', 'Unknown')}\n"
    md_content += f"- **Manual Review Required**: {'Yes' if quality_metrics.get('manual_review_required', False) else 'No'}\n"
    
    # Validation Results
    validation_summary = validation.get('validation_summary', {})
    md_content += "\n## Validation Summary\n\n"
    md_content += f"- **Total Questions**: {validation_summary.get('total_questions', 0)}\n"
    md_content += f"- **Questions Passed**: {validation_summary.get('questions_passed', 0)}\n"
    md_content += f"- **Questions Failed**: {validation_summary.get('questions_failed', 0)}\n"
    md_content += f"- **Questions with Warnings**: {validation_summary.get('questions_with_warnings', 0)}\n"
    
    # Questions requiring review
    review_questions = validation.get('questions_requiring_review', [])
    if review_questions:
        md_content += "\n### Questions Requiring Review\n\n"
        for q in review_questions:
            md_content += f"- **{q.get('question_id', 'Unknown')}** ({q.get('section', 'Unknown')}): {q.get('reason', 'Unknown')}\n"
    
    # Gaps and Comments
    md_content += "\n## Gaps and Internal Comments\n\n"
    for gap_id, gap_data in gaps.items():
        if isinstance(gap_data, dict):
            question = gap_data.get('question', gap_id)
            answer = gap_data.get('answer', 'MANUAL_FIELD')
            
            md_content += f"### {question}\n"
            if answer != 'MANUAL_FIELD':
                md_content += f"{answer}\n"
            else:
                md_content += "*To be completed manually*\n"
            md_content += "\n"
    
    # Footer
    md_content += "\n---\n"
    md_content += "*This report was generated automatically. Please review all low-confidence items manually.*\n"
    
    # Write the markdown file
    try:
        os.makedirs(os.path.dirname(md_path), exist_ok=True)
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(md_content)
        
        print(f"✅ Markdown report written to: {md_path}")
        md_size = os.path.getsize(md_path)
        print(f"📄 Markdown file size: {md_size:,} bytes")
        
        return md_path
        
    except Exception as e:
        print(f"❌ Error writing markdown report: {e}")
        return ""

def create_markdown_report(analysis_data: Dict[str, Any], md_path: str, target_document_path: str, reference_document_path: str) -> str:
    """
    Create a markdown report from analysis data.
    
    Args:
        analysis_data: Complete analysis results
        md_path: Path where markdown should be written
        target_document_path: Path to analyzed document
        reference_document_path: Path to reference document
        
    Returns:
        Path to created markdown file
    """
    return write_markdown_report(analysis_data, md_path, target_document_path, reference_document_path)

def write_spreadsheet_report(analysis_data: Dict[str, Any], output_path: str, target_document_path: str, reference_document_path: str) -> str:
    """
    Write a spreadsheet (Excel/CSV) with questions as headers and answers as data.
    
    Args:
        analysis_data: The complete analysis data
        output_path: Base path for the YAML file (will change extension to .xlsx/.csv)
        target_document_path: Path to analyzed document
        reference_document_path: Path to reference document
        
    Returns:
        Path to the created spreadsheet file
    """
    # Generate spreadsheet file path
    if EXCEL_AVAILABLE:
        spreadsheet_path = output_path.replace('.yaml', '.xlsx').replace('.yml', '.xlsx')
        file_format = 'Excel'
    else:
        spreadsheet_path = output_path.replace('.yaml', '.csv').replace('.yml', '.csv')
        file_format = 'CSV'
    
    try:
        # Extract all questions and answers
        questions_data = []
        
        # Get document information
        doc_info = analysis_data.get('document_information', {})
        for field_id, field_data in doc_info.items():
            if isinstance(field_data, dict):
                questions_data.append({
                    'Section': 'Document Information',
                    'Field_ID': field_id,
                    'Question': field_data.get('question', field_id),
                    'Answer': field_data.get('answer', 'Not specified'),
                    'Answer_Source': 'Deterministic',
                    'Confidence': field_data.get('confidence', 1.0),
                    'Guideline': field_data.get('guideline', ''),
                    'Requires_Review': field_data.get('validation_status', {}).get('requires_review', False)
                })
        
        # Get contract summary  
        contract_summary = analysis_data.get('contract_summary', {})
        for item_id, item_data in contract_summary.items():
            if isinstance(item_data, dict):
                questions_data.append({
                    'Section': 'Contract Summary',
                    'Field_ID': item_id,
                    'Question': item_data.get('question', item_id),
                    'Answer': item_data.get('answer', 'Not specified'),
                    'Answer_Source': 'LLM',
                    'Confidence': item_data.get('confidence', 0),
                    'Guideline': item_data.get('guideline', ''),
                    'Requires_Review': item_data.get('validation_status', {}).get('requires_review', False)
                })
        
        # Get gaps and comments
        gaps = analysis_data.get('gaps_and_comments', {})
        for gap_id, gap_data in gaps.items():
            if isinstance(gap_data, dict):
                src = 'Manual' if gap_data.get('answer', 'MANUAL_FIELD') == 'MANUAL_FIELD' else 'LLM'
                questions_data.append({
                    'Section': 'Gaps and Comments',
                    'Field_ID': gap_id,
                    'Question': gap_data.get('question', gap_id),
                    'Answer': gap_data.get('answer', 'MANUAL_FIELD'),
                    'Answer_Source': src,
                    'Confidence': gap_data.get('confidence', 1.0),
                    'Guideline': gap_data.get('guideline', ''),
                    'Requires_Review': gap_data.get('validation_status', {}).get('requires_review', False)
                })
        
        # Create DataFrame
        df = pd.DataFrame(questions_data)
        
        # Add metadata columns
        exec_summary = analysis_data.get('executive_summary', {})
        df['Document_Name'] = exec_summary.get('document_name', 'Unknown')
        df['Analysis_Date'] = exec_summary.get('analysis_date', 'Unknown')
        df['Risk_Level'] = exec_summary.get('risk_level', 'Unknown')
        df['Quality_Grade'] = exec_summary.get('quality_grade', 'Unknown')
        df['Overall_Recommendation'] = exec_summary.get('overall_recommendation', 'Unknown')

        # Add company metadata if available
        extracted_entities = analysis_data.get('extracted_entities', {})
        company1 = 'Not found'
        company2 = 'Not found'
        if extracted_entities:
            p1 = extracted_entities.get('party1')
            p2 = extracted_entities.get('party2')
            if isinstance(p1, dict) and p1.get('name'):
                company1 = p1.get('name')
            elif isinstance(extracted_entities.get('companies'), list) and extracted_entities['companies']:
                company1 = str(extracted_entities['companies'][0])
            if isinstance(p2, dict) and p2.get('name'):
                company2 = p2.get('name')
            elif isinstance(extracted_entities.get('companies'), list) and len(extracted_entities['companies']) > 1:
                company2 = str(extracted_entities['companies'][1])
        df['Company 1'] = company1
        df['Company 2'] = company2
        
        # Reorder columns for better readability
        column_order = [
            'Document_Name', 'Analysis_Date', 'Section', 'Field_ID', 'Question', 
            'Answer', 'Answer_Source', 'Confidence', 'Requires_Review', 'Risk_Level', 
            'Quality_Grade', 'Overall_Recommendation', 'Guideline',
            'Company 1', 'Company 2'
        ]
        df = df[column_order]

        # Add per-rule columns to the sheet (repeat per row for convenience)
        rule_results = analysis_data.get('rule_compliance_results') or []
        if rule_results:
            for r in rule_results:
                name = r.get('rule_id') or r.get('name') or 'rule'
                status_col = f"Rule:{name} Status"
                rationale_col = f"Rule:{name} Rationale"
                status_val = str(r.get('status', 'unknown'))
                rationale_val = str(r.get('rationale', ''))
                df[status_col] = status_val
                df[rationale_col] = rationale_val[:200] + ('...' if len(rationale_val) > 200 else '')
        
        # Write file
        os.makedirs(os.path.dirname(spreadsheet_path), exist_ok=True)
        
        if EXCEL_AVAILABLE and file_format == 'Excel':
            # Write Excel file with formatting
            with pd.ExcelWriter(spreadsheet_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Contract Analysis')
                
                # Get the workbook and worksheet
                worksheet = writer.sheets['Contract Analysis']
                
                # Apply formatting
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Header formatting
                Font(bold=True)
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                for cell in worksheet[1]:
                    cell.font = Font(color='FFFFFF', bold=True)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

                # Add symbol legend at bottom
                last_row = len(df) + 3
                worksheet.cell(row=last_row, column=1, value="Legend: ★ = LLM Answer; Confidence column shows per-question confidence")
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Highlight rows that require review
                review_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                for row_num, requires_review in enumerate(df['Requires_Review'], start=2):
                    if requires_review:
                        for col_num in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row_num, column=col_num).fill = review_fill
        else:
            # Write CSV file
            df.to_csv(spreadsheet_path, index=False, encoding='utf-8')
        
        print(f"✅ {file_format} spreadsheet written to: {spreadsheet_path}")
        unique_sections = len(set(df['Section'].tolist()))
        print(f"📊 Contains {len(questions_data)} questions across {unique_sections} sections")
        
        return spreadsheet_path
        
    except Exception as e:
        print(f"❌ Error writing {file_format} file to {spreadsheet_path}: {e}")
        return ""

def create_multi_format_outputs(analysis_data: Dict[str, Any], 
                               output_path: str,
                               target_document_path: str,
                               reference_document_path: str,
                               use_template: bool = True) -> Dict[str, str]:
    """
    Create multiple output formats from analysis data.
    
    Args:
        analysis_data: Complete analysis results 
        output_path: Base path for outputs (typically YAML path)
        target_document_path: Path to analyzed document
        reference_document_path: Path to reference document
        use_template: Whether to use template format for Excel
        
    Returns:
        Dictionary with paths to created outputs
    """
    outputs = {}
    
    try:
        # 1. YAML Output (detailed data)
        outputs['yaml'] = output_path
        print(f"📄 YAML analysis: {output_path}")
        
        # 2. Markdown Report (human-readable)
        md_path = output_path.replace('.yaml', '.md').replace('.yml', '.md')
        create_markdown_report(analysis_data, md_path, target_document_path, reference_document_path)
        outputs['markdown'] = md_path
        print(f"📝 Markdown report: {md_path}")
        
        # 3. Excel Spreadsheet (choose format)
        if use_template:
            # Use template format
            excel_path = output_path.replace('.yaml', '_template.xlsx').replace('.yml', '_template.xlsx')
            template_path = create_template_excel(
                analysis_data, excel_path, target_document_path, reference_document_path
            )
            if template_path:
                outputs['excel_template'] = template_path
        else:
            # Use generic format  
            excel_path = output_path.replace('.yaml', '.xlsx').replace('.yml', '.xlsx')
            write_spreadsheet_report(analysis_data, excel_path, target_document_path, reference_document_path)
            outputs['excel_generic'] = excel_path
            print(f"📊 Generic Excel: {excel_path}")
        
        return outputs
        
    except Exception as e:
        print(f"❌ Error creating multi-format outputs: {e}")
        return {}

def write_organized_outputs(state: Dict[str, Any], 
                          run_id: str,
                          target_document_path: str,
                          reference_document_path: str,
                          use_template: bool = True) -> Dict[str, str]:
    """
    Write outputs using the organized directory structure.
    
    Args:
        state: Complete analysis state
        run_id: Run identifier  
        target_document_path: Path to analyzed document
        reference_document_path: Path to reference document
        use_template: Whether to use template format for Excel
        
    Returns:
        Dictionary with paths to created outputs
    """
    try:
        # Setup run directory
        run_dir, yaml_path = output_organizer.create_run_directory(
            run_id, target_document_path, reference_document_path
        )
        
        # Write YAML analysis data
        with open(yaml_path, 'w', encoding='utf-8') as f:
            yaml.dump(state, f, default_flow_style=False, sort_keys=False, allow_unicode=True)
        
        # Create multi-format outputs
        outputs = create_multi_format_outputs(
            state, yaml_path, target_document_path, reference_document_path, use_template
        )
        
        # Update master Excel file
        master_excel_path = output_organizer.update_master_excel(
            state, run_id, target_document_path, reference_document_path, use_template
        )
        if not master_excel_path:
            # Retry once with template mode forced
            master_excel_path = output_organizer.update_master_excel(
                state, run_id, target_document_path, reference_document_path, True
            )
        if master_excel_path:
            outputs['master_excel'] = master_excel_path
        
        # Generate rule-centric compliance Excel if rules are enabled
        rules_path = state.get('rules_path')
        if rules_path and state.get('rule_compliance_results'):
            try:
                from utils.rule_centric_excel_writer import create_rule_compliance_spreadsheet
                compliance_excel_path = os.path.join(run_dir, 'rule_compliance.xlsx')
                compliance_excel = create_rule_compliance_spreadsheet(
                    state, 
                    compliance_excel_path,
                    target_document_path,
                    rules_path
                )
                if compliance_excel:
                    outputs['rule_compliance_excel'] = compliance_excel
                    print("   📊 Rule Compliance Excel: rule_compliance.xlsx")
                
                # Also update master compliance workbook for batch processing
                from utils.rule_compliance_master_writer import update_master_compliance_excel
                master_compliance_path = os.path.join(
                    os.path.dirname(run_dir), 
                    'comparisons', 
                    'rule_compliance_master.xlsx'
                )
                master_compliance = update_master_compliance_excel(
                    state,
                    target_document_path,
                    rules_path,
                    master_compliance_path
                )
                if master_compliance:
                    outputs['master_compliance_excel'] = master_compliance
                    print("   📊 Master Compliance: rule_compliance_master.xlsx")
            except Exception as e:
                print(f"   ⚠️ Could not generate rule compliance Excel: {e}")
        
        # Create summary dashboard
        dashboard_path = output_organizer.create_summary_dashboard()
        if dashboard_path:
            outputs['dashboard'] = dashboard_path
        
        # Archive any old flat-structure outputs
        output_organizer.archive_old_outputs()
        
        print("\n✅ Organized outputs complete:")
        print(f"   📁 Run directory: {run_dir}")
        print("   📄 Analysis data: analysis.yaml")
        print("   📝 Report: analysis.md")
        if use_template:
            print("   📊 Template Excel: analysis_template.xlsx")
        else:
            print("   📊 Generic Excel: analysis.xlsx")
        print(f"   🔗 Master comparison: {os.path.basename(master_excel_path) if master_excel_path else 'Failed'}")
        
        return outputs
        
    except Exception as e:
        print(f"❌ Error writing organized outputs: {e}")
        return {}

def questionnaire_yaml_populator_node(state: Dict[str, Any]) -> Dict[str, Any]:
    """
    Populate questionnaire, create analysis outputs, and organize files in structured directories.
    Enhanced to support template format.
    """
    print("\n" + "="*80)
    print("📋 QUESTIONNAIRE POPULATOR - Template Support")
    print("="*80)
    
    try:
        # Extract paths
        target_document_path = state.get('target_document_path', '')
        reference_document_path = state.get('reference_document_path', '')
        
        # Check for template format preference (default: True)
        use_template = state.get('use_template', True)
        
        if not target_document_path:
            raise ValueError("target_document_path not found in state")
        
        print(f"📄 Target document: {os.path.basename(target_document_path)}")
        print(f"📚 Reference document: {os.path.basename(reference_document_path) if reference_document_path else 'None'}")
        print(f"📊 Excel format: {'Template' if use_template else 'Generic'}")
        
        # Load questionnaire template (guarded)
        questionnaire_path = os.getenv('QUESTIONNAIRE_PATH', "questionnaires/contract_evaluation.yaml")
        try:
            if not os.path.exists(questionnaire_path):
                print(f"⚠️ Questionnaire template not found: {questionnaire_path}. Continuing without it.")
            else:
                with open(questionnaire_path, 'r', encoding='utf-8') as f:
                    yaml.safe_load(f)
                    print(f"📋 Loaded questionnaire template: {questionnaire_path}")
        except Exception as qe:
            print(f"⚠️ Failed to load questionnaire template: {qe}. Proceeding without template.")
        
        # Get deterministic fields first
        deterministic_fields = extract_deterministic_fields(state)  # type: ignore
        
        # Get questionnaire responses
        questionnaire_responses = state.get('questionnaire_responses', {})
        
        # Calculate quality metrics
        quality_report = quality_analyzer.calculate_overall_quality_rating(dict(state))
        
        # Generate risk summary
        risk_summary = create_risk_summary(state)  # type: ignore
        
        # Get red flag analysis
        classified_sentences = state.get('classified_sentences', [])
        dict_sentences = [dict(sentence) for sentence in classified_sentences]
        red_flag_analysis = red_flag_detector.analyze_document_red_flags(dict_sentences)
        
        # Calculate processing statistics
        processing_stats = calculate_processing_stats(state)  # type: ignore
        
        # Create executive summary
        executive_summary = create_executive_summary(state, quality_report, risk_summary, red_flag_analysis)  # type: ignore
        
        # Extract questionnaire data with enhanced validation
        questionnaire_data = extract_questionnaire_data(questionnaire_responses, deterministic_fields)
        
        # Create validation summary
        validation_summary = {
            'total_questions': questionnaire_data['validation_results']['questions_processed'],
            'questions_passed': (
                questionnaire_data['validation_results']['questions_processed'] -
                questionnaire_data['validation_results']['questions_with_errors'] -
                questionnaire_data['validation_results']['questions_with_low_confidence']
            ),
            'questions_failed': questionnaire_data['validation_results']['questions_with_errors'],
            'questions_with_warnings': questionnaire_data['validation_results']['questions_with_low_confidence']
        }
        
        # Compile final analysis data
        analysis_data = {
            'executive_summary': executive_summary,
            'document_information': questionnaire_data['document_information'],
            'contract_summary': questionnaire_data['contract_summary'],
            'gaps_and_comments': questionnaire_data['gaps_and_comments'],
            'validation_results': {
                **questionnaire_data['validation_results'],
                'overall_validation_status': 'PASSED' if (
                    questionnaire_data['validation_results']['questions_with_errors'] == 0 and
                    questionnaire_data['validation_results']['questions_with_low_confidence'] < 3
                ) else 'FAILED',
                'validation_summary': validation_summary
            },
            'processing_metadata': {
                **processing_stats,
                'error_summary': questionnaire_data['processing_metadata']['error_summary']
            },
            'quality_metrics': quality_report,
            'risk_assessment': {
                **risk_summary,
                'red_flags': red_flag_analysis.get('red_flags', []),
                'risk_distribution': {
                    'critical': len([f for f in red_flag_analysis.get('red_flags', []) if f.get('severity') == 'CRITICAL']),
                    'high': len([f for f in red_flag_analysis.get('red_flags', []) if f.get('severity') == 'HIGH']),
                    'medium': len([f for f in red_flag_analysis.get('red_flags', []) if f.get('severity') == 'MEDIUM']),
                    'low': len([f for f in red_flag_analysis.get('red_flags', []) if f.get('severity') == 'LOW'])
                }
            }
        }

        # If rules results present on state, ensure they are included in analysis for Excel writers
        # Always include compliance summary and rules execution status, even if results are empty
        rule_results = state.get('rule_compliance_results')
        rules_mode_state = state.get('rules_mode') or {}
        # Prefer explicit rules_data count; fall back to number of results if data not persisted
        rules_count = len(state.get('rules_data') or [])
        if rules_count == 0 and state.get('rule_compliance_results'):
            rules_count = len(state.get('rule_compliance_results') or [])
        analysis_data['rule_compliance_results'] = rule_results or []
        if rule_results:
            counts = {"compliant": 0, "non_compliant": 0, "not_applicable": 0, "unknown": 0}
            for r in rule_results:
                status = str(r.get('status', 'unknown')).lower()
                counts[status] = counts.get(status, 0) + 1
            analysis_data['compliance_summary'] = {
                'status_counts': counts,
                'total_rules': len(rule_results),
            }
        else:
            analysis_data['compliance_summary'] = {
                'status': 'unknown' if rules_count > 0 else 'not_evaluated',
                'message': f"Rules present: {rules_count}. {'No results produced.' if rules_count>0 else 'Rules not loaded.'}",
            }
        # Surface rules mode execution flags in output
        analysis_data['rules_mode'] = {
            'enabled': bool(rules_mode_state.get('enabled', True)),
            'executed': bool(rules_mode_state.get('executed', True)),
            'retrieval_fallback_used': bool(rules_mode_state.get('retrieval_fallback_used', False)),
            'rules_loaded_count': rules_count,
        }
        # Pass through extracted entities for Company 1/2 columns
        if state.get('extracted_entities') is not None:
            analysis_data['extracted_entities'] = state.get('extracted_entities')
        
        # Generate run ID and create outputs
        run_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Create organized outputs using the new system
        # Respect environment override to persist master template across runs
        env_use_template = os.getenv('USE_TEMPLATE_EXCEL', 'false').lower() in ('1', 'true', 'yes', 'on')
        outputs = write_organized_outputs(
            analysis_data, run_id, target_document_path, reference_document_path, use_template or env_use_template
        )
        
        print(f"\n🎯 Analysis complete with {'template' if use_template else 'generic'} format")
        print(f"📊 Quality: {quality_report.get('quality_grade', 'Unknown')} ({quality_report.get('overall_quality_score', 0):.0f}%)")
        print(f"🚨 Risk Level: {executive_summary.get('risk_level', 'Unknown')}")
        print(f"📋 Validation: {validation_summary['questions_passed']}/{validation_summary['total_questions']} questions passed")
        
        # Update state with analysis results
        state.update({
            'analysis_complete': True,
            'analysis_data': analysis_data,
            'outputs': outputs,
            'run_id': run_id,
            'use_template': use_template
        })
        
        return state
        
    except Exception as e:
        print(f"❌ Error in questionnaire_yaml_populator_node: {e}")
        state['errors'] = state.get('errors', []) + [str(e)]
        return state